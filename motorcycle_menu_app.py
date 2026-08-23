# -*- coding: utf-8 -*-
"""
برنامه انتخاب منوی تشخیص موتورسیکلت (نسخه دسکتاپ)
=====================================================
این برنامه همون منطق کرکره‌ای درختی شیت Ecu_Menu_Form اکسل رو
به صورت یک برنامه‌ی مستقل ویندوزی (بدون نیاز به اکسل) پیاده می‌کنه:

    برند  ->  مدل (وابسته به برند)  ->  واحد (وابسته به برند+مدل)
    -> جدول آیتم‌های منو (خودکار پر می‌شه)

قابلیت‌ها:
  - با تغییر برند، مدل و واحد و جدول آیتم‌ها خودکار پاک می‌شن.
  - با تغییر مدل، واحد و جدول آیتم‌ها خودکار پاک می‌شن.
  - بعد از هر انتخاب، فوکوس و کرکره‌ی مرحله‌ی بعد خودکار باز می‌شه.
  - تمام متن‌های جدول قابل انتخاب و کپی هستن (کلیک+درگ یا Ctrl+A سپس Ctrl+C).
  - دکمه «افزودن به لیست خروجی» برای جمع کردن چند ردیف در یک لیست.
  - دکمه «خروجی اکسل» برای ذخیره‌ی لیست به یک فایل .xlsx (با رنگ‌بندی مثل نمونه اصلی).
  - ذخیره/بارگذاری پروژه (JSON) برای ادامه‌ی کار در جلسه‌ی بعد.
  - اتصال مستقیم به فایل اکسل دیتابیس (شیت‌های Vehicles/Option/Unit/Ecu_Menu):
    از دکمه‌ی «🗄 دیتابیس» مسیر فایل اکسل رو یک‌بار مشخص می‌کنید، برنامه اون مسیر
    رو ذخیره می‌کنه، و هر وقت اکسل عوض بشه کافیه دکمه‌ی «🔄 بارگذاری مجدد داده‌ها»
    رو بزنید تا دیتای برنامه به‌روز بشه (نیازی به هیچ تغییر دستی کد نیست).

نیازمندی‌ها: پایتون ۳.9+ ، پکیج openpyxl (برای خواندن/نوشتن اکسل)
    pip install openpyxl

اگر فایل اکسل هنوز تنظیم نشده باشه، برنامه به‌صورت پیش‌فرض از tree_data.json
(در صورت وجود کنار همین فایل) استفاده می‌کنه.
"""

import json
import os
import re
import sys
import hashlib
import threading
import tempfile
import urllib.request
import urllib.error
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# اگه برنامه با python.exe معمولی اجرا شده باشه (نه exe ساخته‌شده با
# --windowed یا pythonw)، یک پنجره‌ی سیاه cmd پشت برنامه باز می‌مونه.
# این چند خط، در صورت وجود چنین پنجره‌ای، همون لحظه‌ی شروع مخفیش می‌کنه.
if sys.platform == "win32":
    try:
        import ctypes
        _console_hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if _console_hwnd:
            ctypes.windll.user32.ShowWindow(_console_hwnd, 0)  # SW_HIDE
    except Exception:
        pass
    try:
        import ctypes
        # DPI-awareness رو فعال می‌کنیم؛ وگرنه روی مانیتورهایی که مقیاسِ
        # ویندوز (Windows Scaling) بیشتر از ۱۰۰٪ باشه، مختصاتِ ویجت‌ها
        # (که برای گرفتنِ عکسِ لحظه‌ای در افکتِ مورف استفاده می‌شن) با
        # پیکسل‌های واقعیِ صفحه هم‌خوانی نداره و عکس‌ها کج/جابه‌جا می‌شن.
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

APP_TITLE = "انتخاب منوی تشخیص موتورسیکلت"
DATA_FILE = "tree_data.json"
CONFIG_FILE = "app_settings.json"
REQUIRED_SHEETS = ["Vehicles", "Option", "Unit", "Ecu_Menu"]
# آدرس ثابت ریپوی فایل‌های اکسل روی گیت‌هاب (قابل تغییر توسط کاربر نیست)
GITHUB_EXCEL_RAW_BASE = "https://raw.githubusercontent.com/workrezvanian-netizen/DiagExcelFiles/main"
# وضعیت فایل‌های دانلود‌شده: rel_path -> {etag, size, sha256}
CFG_LOCAL_EXCEL_META = "local_excel_meta"

# کش محلی فایل‌های دانلود‌شده از گیت‌هاب (دیگر از پوشهٔ اکسل کاربر خوانده نمی‌شود)
CACHE_DIR_NAME = "excel_cache"
MOTORCYCLE_EXCEL_NAME = "18_MotorCycle.xlsm"

# فایل‌های اکسلِ مربوط به شرکت‌های خودروساز، به همراه اسم فارسیِ نمایشی‌شون
# (به همون ترتیبی که توی پوشه‌ی دیاگ شماره‌گذاری شدن).
CAR_MAKER_FILES = [
    ("01_IranKhodro.xlsm", "ایران خودرو"),
    ("02_Saipa.xlsm", "سایپا"),
    ("03_Zamyad.xlsm", "زامیاد"),
    ("04_Parskhodro.xlsm", "پارس خودرو"),
    ("05_ModiranKhodro.xlsm", "مدیران خودرو"),
    ("06_KermanMotor.xlsm", "کرمان موتور"),
    ("07_Bahman.xlsm", "بهمن موتور"),
    ("08_Geely.xlsm", "جیلی"),
    ("09_Kia.xlsm", "کیا"),
    ("10_Hyundai.xlsm", "هیوندای"),
    ("11_KMC.xlsm", "KMC"),
    ("12_MG.xlsm", "ام‌جی"),
    ("13_Lamari.xlsm", "لاماری"),
    ("14_BYD.xlsm", "BYD"),
    ("15_RENAULT.xlsm", "رنو"),
    ("16_GREATWALL.xlsm", "گریت وال"),
    ("17_BAIC.xlsm", "BAIC"),
]
CAR_MAKER_LABELS = [label for _fname, label in CAR_MAKER_FILES]
_CAR_LABEL_TO_FILE = {label: fname for fname, label in CAR_MAKER_FILES}

ITEM_ROWS = 6

COLOR_HEADER = "#E7E6E6"
COLOR_VEHICLE = "#E26E79"
COLOR_MODEL = "#9BC2E6"
COLOR_UNIT = "#FFE699"
COLOR_SUBUNIT = "#C6E0B4"
COLOR_ITEM = "#FEE2F6"
COLOR_BG = "#FFFFFF"

# رنگ‌های «صفحه‌ی دستگاه» (شبیه صفحه‌نمایش یک دستگاه تشخیص/اسکنر واقعی)
# رنگ‌های «صفحه‌ی دستگاه» — الهام‌گرفته از صفحه‌نمایش دستگاه‌های
# تشخیص واقعی (نوار بالای آبی، لیست با پس‌زمینه‌ی سفید، ردیف انتخاب‌شده‌ی
# فیروزه‌ای/آبی‌روشن)
DEV_BEZEL = "#1b1b1b"
DEV_HEADER_BG = "#1f5fb8"
DEV_HEADER_TEXT = "#ffffff"
DEV_HEADER_TEXT_DIM = "#bcd4f5"
DEV_HEADER_HOVER = "#2f77d8"
DEV_BODY_BG = "#ffffff"
DEV_ROW_BORDER = "#e2e6ea"
DEV_SELECT_BG = "#29c4e6"
DEV_SELECT_FG = "#ffffff"
DEV_TEXT = "#202020"
DEV_BULLET = "#d98c3d"
# نام‌های قدیمی برای سازگاری با بقیه‌ی کد
DEV_SCREEN = DEV_BODY_BG
DEV_SCREEN_BORDER = DEV_HEADER_BG
DEV_ACCENT = DEV_HEADER_BG
DEV_BTN_BG = DEV_HEADER_BG
DEV_BTN_HOVER = DEV_HEADER_HOVER
DEV_TEXT_DIM = "#8a8f98"

ICONS_DIR = "brand_icons"
ICON_SIZE = 60


# متن ترکیبیِ فارسی+انگلیسی (مثلاً «کاوان S» یا «بوش یورو3») توی این محیط
# مشکل داره: با عکس‌هایی که فرستادی و مقایسه‌شون با متن خامِ خودِ اکسل،
# دقیق مشخص شد Tk موقع نمایش، ترتیبِ تکه‌های مختلفِ زبان (فارسی در برابر
# انگلیسی/عدد) رو نسبت به هم برعکس نمی‌کنه — انگار همون ترتیبِ ذخیره‌شده،
# تکه به تکه، کنار هم چیده می‌شه و کل بلوک راست‌چین می‌شه (هرچند داخل
# خودِ هر تکه‌ی فارسی، ترتیب کلمات درست چیده می‌شه). راه‌حل: خودمون از
# قبل ترتیبِ تکه‌ها (نه محتوای داخلشون) رو برعکس می‌کنیم، طوری که وقتی
# Tk دوباره (بدون برعکس‌کردن) کنار هم بذارتشون، نتیجه‌ی نهایی درست بشه.
_PERSIAN_CHAR_RE = re.compile(r"[\u0600-\u06FF]")
_BIDI_RUN_RE = re.compile(r"[\u0600-\u06FF]+(?:[ \u200c]+[\u0600-\u06FF]+)*|[^\u0600-\u06FF]+")
_BIDI_ADJ_GAP_RE = re.compile(r"(?<=[A-Za-z0-9])(?=[\u0600-\u06FF])|(?<=[\u0600-\u06FF])(?=[A-Za-z0-9])")


def bidi_fix(text):
    if not text or not _PERSIAN_CHAR_RE.search(text):
        return text  # فارسی نداره؛ کاری لازم نیست
    runs = _BIDI_RUN_RE.findall(text)
    if len(runs) <= 1:
        return text  # فقط یک تکه‌ست (همه فارسی، یا همه غیرفارسی)؛ چیزی برای جابه‌جایی نیست
    joined = "".join(reversed(runs)).strip()
    return _BIDI_ADJ_GAP_RE.sub(" ", joined)


def is_persian_text(text):
    """True اگه رشته حداقل یک حرف فارسی/عربی داشته باشه. برای انتخاب فونتِ
    درست استفاده می‌شه: متنِ کاملاً انگلیسی (مثل «MSE60») باید با فونت
    انگلیسی نمایش داده بشه، وگرنه فونت فارسی حتی عددهاش رو هم به شکل
    فارسی (۶۰ به‌جای 60) نشون می‌ده."""
    return bool(text) and bool(_PERSIAN_CHAR_RE.search(text))


def split_text_runs(text):
    """متن رو به تکه‌های همگن فارسی / غیرفارسی (لاتین+عدد) می‌شکنه. برای
    رشته‌های ترکیبی (مثل «دلفی یورو 2» یا «سنسور MAP») لازمه، چون اگه کل
    رشته با یک فونت فارسی نمایش داده بشه، حتی بخش‌های عددی/لاتینش هم به
    سبک فارسی (مثلاً ۲ به‌جای 2) کشیده می‌شن. با این تابع هر تکه بعداً با
    فونت خودش (فارسی یا انگلیسی) جدا رندر می‌شه تا اعداد/لاتین همیشه به
    شکل انگلیسیِ عادی بمونن."""
    if not text:
        return []
    return [r.strip() for r in _BIDI_RUN_RE.findall(text) if r.strip()]

# فونت‌های زیبا برای نمایش فارسی و انگلیسی.
# لیست اولویت: اگر فونت اول روی سیستم کاربر نصب نباشد، به فونت بعدی می‌رویم
# (این‌طوری همیشه یک فونت خوانا و زیبا تضمین می‌شود، نه یک فونت پیش‌فرض زشت).
# «نازنین» با چند نام رایج نصب (روی ویندوزهای مختلف با نام‌های کمی متفاوت
# نصب می‌شود) اول اولویت است.
FONT_PREFERENCE = ["Nazanin", "B Nazanin", "IRNazanin", "XNazanin",
                    "B Titr", "Vazirmatn", "IRANSans", "Tahoma", "Arial"]
# فونت زیبا برای متن‌های خالص انگلیسی (دکمه‌ها، ستون English/Old ID و ...)
FONT_EN_PREFERENCE = ["Segoe UI", "Calibri", "Century Gothic", "Verdana", "Arial"]


def pick_available_font(preference, fallback):
    try:
        import tkinter.font as tkfont
        installed = set(tkfont.families())
    except Exception:
        installed = set()
    for name in preference:
        if name in installed:
            return name
    return fallback  # همیشه روی ویندوز موجود است


FONT_FA = "Tahoma"  # مقدار موقت؛ در ابتدای اجرای برنامه با pick_available_font() به‌روزرسانی می‌شود
FONT_EN_NAME = "Segoe UI"
FONT_TITLE = None
FONT_LABEL = None
FONT_COMBO = None
FONT_HEADER = None
FONT_CELL_BOLD = None
FONT_CELL = None
FONT_CELL_EN = None
FONT_CELL_EN_BOLD = None
FONT_BUTTON = None
FONT_BUTTON_EN = None
FONT_GROUP = None
FONT_SCREEN_ROW = None
FONT_SCREEN_ROW_EN = None


def init_fonts():
    """باید فقط بعد از ساخته‌شدن پنجره‌ی اصلی Tk صدا زده شود (برای دسترسی به لیست فونت‌های سیستم)."""
    global FONT_FA, FONT_EN_NAME, FONT_TITLE, FONT_LABEL, FONT_COMBO, FONT_HEADER
    global FONT_CELL_BOLD, FONT_CELL, FONT_CELL_EN, FONT_CELL_EN_BOLD
    global FONT_BUTTON, FONT_BUTTON_EN, FONT_GROUP, FONT_SCREEN_ROW, FONT_SCREEN_ROW_EN
    FONT_FA = pick_available_font(FONT_PREFERENCE, "Tahoma")
    FONT_EN_NAME = pick_available_font(FONT_EN_PREFERENCE, "Segoe UI")
    FONT_TITLE = (FONT_FA, 15, "bold")
    FONT_LABEL = (FONT_FA, 10, "bold")
    FONT_COMBO = (FONT_FA, 10)
    FONT_HEADER = (FONT_FA, 10, "bold")
    FONT_CELL_BOLD = (FONT_FA, 10, "bold")
    FONT_CELL = (FONT_FA, 9)
    FONT_CELL_EN = (FONT_EN_NAME, 9)
    FONT_CELL_EN_BOLD = (FONT_EN_NAME, 9, "bold")
    FONT_BUTTON = (FONT_FA, 9, "bold")
    FONT_BUTTON_EN = (FONT_EN_NAME, 9, "bold")
    FONT_GROUP = (FONT_FA, 9, "bold")
    FONT_SCREEN_ROW = (FONT_FA, 13, "bold")
    # فونتِ انگلیسی هم‌سایز، برای وقتی متنِ یک ردیف کاملاً انگلیسیه (مثل
    # «MSE60»، «EPM44») — چون فونت فارسی حتی برای عددهای انگلیسیِ داخل
    # چنین متن‌هایی گلیفِ فارسی (۴۴ به‌جای 44) نشون می‌ده.
    FONT_SCREEN_ROW_EN = (FONT_EN_NAME, 13, "bold")


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def app_dir():
    """پوشه‌ای که فایل تنظیمات (app_settings.json) توش ذخیره می‌شه.
    وقتی به exe تبدیل شده، کنار خود فایل exe؛ در غیر این صورت کنار اسکریپت پایتون."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def config_path():
    return os.path.join(app_dir(), CONFIG_FILE)


def load_config():
    path = config_path()
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(cfg):
    try:
        with open(config_path(), "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def excel_cache_root():
    """ریشهٔ کش محلی اکسل‌های دانلود‌شده از گیت‌هاب."""
    root = os.path.join(app_dir(), CACHE_DIR_NAME)
    os.makedirs(root, exist_ok=True)
    return root


def diag_cache_folder():
    path = os.path.join(excel_cache_root(), "Diag_Menu")
    os.makedirs(path, exist_ok=True)
    return path


def database_cache_folder():
    path = os.path.join(excel_cache_root(), "Diag_Database")
    os.makedirs(path, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# به‌روزرسانی فایل‌های اکسل از گیت‌هاب (بدون مانیفست)
# ---------------------------------------------------------------------------
# فقط فایل‌های شناخته‌شده در Diag_Menu روی ریپو بررسی می‌شوند.
# مقایسه با ETag / اندازه / هش محلی انجام می‌شود.
# ---------------------------------------------------------------------------

def _file_sha256(path, chunk_size=1024 * 1024):
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


def _http_get_bytes(url, token=None, timeout=60):
    req = urllib.request.Request(url, method="GET")
    req.add_header("User-Agent", "MotorcycleMenuSelector/1.0")
    if token:
        req.add_header("Authorization", f"token {token}")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _http_head_meta(url, token=None, timeout=20):
    """ETag و Content-Length را با درخواست HEAD می‌گیرد. اگر HEAD پشتیبانی نشود None."""
    req = urllib.request.Request(url, method="HEAD")
    req.add_header("User-Agent", "MotorcycleMenuSelector/1.0")
    if token:
        req.add_header("Authorization", f"token {token}")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            etag = (resp.headers.get("ETag") or "").strip().strip('"')
            size = resp.headers.get("Content-Length")
            try:
                size = int(size) if size is not None else None
            except ValueError:
                size = None
            return {"etag": etag, "size": size, "status": getattr(resp, "status", 200)}
    except Exception:
        return None


def known_excel_rel_paths():
    """لیست مسیرهای نسبی فایل‌های اکسل روی ریپوی گیت‌هاب."""
    paths = [f"Diag_Menu/{MOTORCYCLE_EXCEL_NAME}"]
    for fname, _label in CAR_MAKER_FILES:
        paths.append(f"Diag_Menu/{fname}")
    return paths


def find_updates_without_manifest(base_url, local_meta, diag_folder, token=None, force_all=False):
    """
    بدون مانیفست: برای هر فایل شناخته‌شده، وجود محلی + ETag/اندازه سرور را چک می‌کند.
    force_all=True → همهٔ فایل‌های موجود روی سرور را دوباره دانلود می‌کند.
    """
    base = (base_url or "").rstrip("/")
    updates = []
    local_meta = local_meta or {}

    for rel_path in known_excel_rel_paths():
        name = rel_path.split("/", 1)[-1]
        local_path = os.path.join(diag_folder, name)
        url = f"{base}/{rel_path}"
        meta = _http_head_meta(url, token=token)

        if meta is not None and meta.get("status") == 404:
            continue

        remote_etag = (meta or {}).get("etag") or ""
        remote_size = (meta or {}).get("size")

        if not os.path.isfile(local_path) or force_all:
            updates.append({
                "rel_path": rel_path,
                "local_path": local_path,
                "download_name": name,
                "remote_etag": remote_etag,
                "remote_size": remote_size,
            })
            continue

        if meta is None:
            continue

        stored = local_meta.get(rel_path) or {}
        local_etag = stored.get("etag") or ""
        try:
            disk_size = os.path.getsize(local_path)
        except OSError:
            disk_size = None

        # ETag فرق کند (حتی اگر قبلاً ذخیره نشده)، یا اندازه فرق کند
        needs = False
        if remote_etag and remote_etag != local_etag:
            needs = True
        elif remote_size is not None and disk_size is not None and int(remote_size) != int(disk_size):
            needs = True

        if needs:
            updates.append({
                "rel_path": rel_path,
                "local_path": local_path,
                "download_name": name,
                "remote_etag": remote_etag,
                "remote_size": remote_size,
            })
    return updates


def download_excel_file(base_url, rel_path, dest_path, token=None, timeout=120):
    """
    یک فایل اکسل را از raw گیت‌هاب دانلود و در dest_path ذخیره می‌کند
    (اول در فایل موقت، بعد جابه‌جایی اتمی برای جلوگیری از خراب شدن فایل).
    """
    import time
    base = (base_url or "").rstrip("/")
    # پارامتر زمان برای دور زدن کش CDN گیت‌هاب
    url = f"{base}/{rel_path.replace(chr(92), '/')}?t={int(time.time())}"
    data = _http_get_bytes(url, token=token, timeout=timeout)
    os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        suffix=".tmp",
        prefix="excel_upd_",
        dir=os.path.dirname(dest_path) or None)
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        # روی ویندوز ممکن است فایل مقصد باز باشد؛ سعی می‌کنیم جایگزین کنیم
        if os.path.exists(dest_path):
            try:
                os.replace(tmp, dest_path)
            except PermissionError:
                # اگر فایل قفل باشد، با نام جدید ذخیره و بعد تلاش برای جایگزینی
                bak = dest_path + ".new"
                os.replace(tmp, bak)
                try:
                    os.replace(bak, dest_path)
                except Exception:
                    # حداقل فایل .new کنارش هست
                    raise PermissionError(
                        f"فایل در حال استفاده است؛ نسخهٔ جدید ذخیره شد در:\n{bak}")
        else:
            os.replace(tmp, dest_path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass
    return _file_sha256(dest_path)


# ---------------------------------------------------------------------------
# خواندن مستقیم دیتای درختی (برند > مدل > واحد > آیتم) از روی فایل اکسل دیتابیس.
# ساختار مورد انتظار در هر شیت: چند «بلوک» کوچک که هرکدوم با ردیف تیتر
# «ردیف | متن فارسی | English | Old ID [| پیغام...]» شروع می‌شن، بعد چند
# ردیفِ «مسیر» (برند/مدل/واحد) و در آخر ردیف‌های داده. این ساختار همون چیزیه
# که در شیت‌های Vehicles / Option / Unit / Ecu_Menu فایل اکسل استفاده شده.
# ---------------------------------------------------------------------------

# بعضی فایل‌های اکسل (مثلاً کیا/هیوندای/لاماری) به‌جای حروف فارسیِ «ی» و
# «ک»، از معادل عربیِ «ي» و «ك» استفاده کردن (چون با کیبورد/لوکیل عربی
# تایپ شدن). این دو تا از نظر یونیکد کاراکترهای متفاوتی‌ان، پس هم باعث می‌شه
# سرستون «ردیف» شناسایی نشه (و کل فایل خالی به نظر برسه)، هم ممکنه نمایش
# بعضی کلمات کمی به‌هم‌ریخته/متفاوت بشه. همه‌جا این‌ها رو به معادل فارسی
# استاندارد تبدیل می‌کنیم.
_ARABIC_TO_PERSIAN_MAP = str.maketrans({"ي": "ی", "ك": "ک"})


def _xl_clean(v):
    if v is None:
        return ""
    return str(v).strip().translate(_ARABIC_TO_PERSIAN_MAP)


def _xl_find_anchors(ws):
    """پیدا کردن گوشه‌ی بالا-چپ همه‌ی بلوک‌های کوچیک توی یک شیت
    (هر جا که سلولش، بعد از یکسان‌سازیِ حروف عربی/فارسی، برابر «ردیف» باشه)."""
    anchors = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and _xl_clean(cell.value) == "ردیف":
                anchors.append((cell.row, cell.column))
    return anchors


def _xl_read_block(ws, r0, c0):
    """از روی گوشه‌ی یک بلوک (r0,c0)، ردیف‌های «مسیر» (برند/مدل/واحد/...)
    رو تا رسیدن به اولین ردیفِ داده‌ی واقعی (ردیفی که ستون «ردیف» توش عدد
    داره) می‌خونه، بعد خودِ ردیف‌های داده رو تا رسیدن به یک ردیف خالی.
    عمق مسیر پویاست: بعضی بلوک‌ها ۳ سطح مسیر دارن (برند/مدل/واحد) و بعضی‌ها
    (وقتی یک واحد چند زیرمجموعه/نسخه داره، مثل «دلفی» با یورو۲/یورو۵) ۴
    سطح (برند/مدل/واحد/زیرواحد) — این تابع خودش این تفاوت رو تشخیص می‌ده."""
    fa_col, en_col, oid_col = c0 + 1, c0 + 2, c0 + 3
    path_levels = []
    row = r0 + 1
    while True:
        idx_val = ws.cell(row, c0).value
        fa_val = ws.cell(row, fa_col).value
        if fa_val is None or str(fa_val).strip() == "":
            break
        if idx_val is None or str(idx_val).strip() == "":
            en = _xl_clean(ws.cell(row, en_col).value)
            oid = _xl_clean(ws.cell(row, oid_col).value)
            path_levels.append((_xl_clean(fa_val), en, oid))
            row += 1
            continue
        break
    items = []
    while True:
        fa_val = ws.cell(row, fa_col).value
        if fa_val is None or str(fa_val).strip() == "":
            break
        fa = _xl_clean(fa_val)
        en = _xl_clean(ws.cell(row, en_col).value)
        oid = _xl_clean(ws.cell(row, oid_col).value)
        items.append((fa, en, oid))
        row += 1
    return path_levels, items


def _ensure_openpyxl(parent=None):
    """مطمئن می‌شه پکیج openpyxl روی سیستم موجوده؛ اگه نبود، تلاش می‌کنه
    خودش (فقط وقتی برنامه به‌صورت فایل .py با پایتون اجرا شده باشه، نه
    به‌صورت exe) به‌طور خودکار نصبش کنه. برای exe، راهنمایی می‌کنه که
    build_exe.bat دوباره اجرا بشه."""
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        pass

    if getattr(sys, "frozen", False):
        messagebox.showerror(
            "پکیج openpyxl نصب نیست",
            "این فایل اجرایی (exe) بدون پکیج openpyxl ساخته شده و نمی‌تونه\n"
            "فایل اکسل رو بخونه یا بنویسه.\n\n"
            "لطفاً فایل build_exe.bat رو دوباره اجرا کنید تا نسخه‌ی exe\n"
            "جدید، این‌بار همراه با openpyxl، ساخته بشه.",
            parent=parent)
        return False

    resp = messagebox.askyesno(
        "نصب پکیج مورد نیاز",
        "برای کار با فایل اکسل، برنامه به پکیج «openpyxl» نیاز داره که روی\n"
        "این سیستم نصب نیست.\n\n"
        "می‌خواید همین الان به‌صورت خودکار نصب بشه؟\n"
        "(نیاز به اتصال اینترنت داره و چند لحظه طول می‌کشه)",
        parent=parent)
    if not resp:
        return False

    try:
        import subprocess
        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
                               creationflags=creationflags)
        import importlib
        importlib.invalidate_caches()
        import openpyxl  # noqa: F401
        messagebox.showinfo("موفق", "پکیج openpyxl با موفقیت نصب شد.", parent=parent)
        return True
    except Exception as e:
        messagebox.showerror(
            "خطا در نصب خودکار",
            "نصب خودکار موفق نبود. لطفاً این دستور رو دستی توی Command Prompt\n"
            "(یا PowerShell) اجرا کنید و دوباره برنامه رو باز کنید:\n\n"
            "    pip install openpyxl\n\n"
            f"جزئیات فنی خطا:\n{e}",
            parent=parent)
        return False


def build_tree_from_excel(path):
    """دیتای درختی برنامه رو مستقیماً از فایل اکسل دیتابیس (بدون نیاز به
    tree_data.json) می‌سازه. هر بار که این تابع صدا زده بشه، آخرین نسخه‌ی
    فایل اکسل خونده می‌شه؛ یعنی با تغییر اکسل، فقط کافیه دوباره صداش بزنیم."""
    if not _ensure_openpyxl():
        raise RuntimeError("برای خواندن فایل اکسل، پکیج openpyxl لازمه و نصب/فعال نشد.")
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(
            "شیت‌های زیر توی فایل اکسل پیدا نشدن:\n" + "، ".join(missing) +
            "\n\nشیت‌های لازم: " + "، ".join(REQUIRED_SHEETS)
        )

    vehicles = []
    ws = wb["Vehicles"]
    for r, c in _xl_find_anchors(ws):
        _, items = _xl_read_block(ws, r, c)
        for fa, en, _oid in items:
            vehicles.append([fa, en])

    models_by_veh = {}
    ws = wb["Option"]
    for r, c in _xl_find_anchors(ws):
        path_levels, items = _xl_read_block(ws, r, c)
        if not path_levels:
            continue
        vfa = path_levels[0][0]
        models_by_veh.setdefault(vfa, [])
        for fa, en, _oid in items:
            models_by_veh[vfa].append([fa, en])

    units_by_vm = {}
    ws = wb["Unit"]
    for r, c in _xl_find_anchors(ws):
        path_levels, items = _xl_read_block(ws, r, c)
        if len(path_levels) < 2:
            continue
        vfa, mfa = path_levels[0][0], path_levels[1][0]
        key = f"{vfa}||{mfa}"
        units_by_vm.setdefault(key, [])
        for fa, en, _oid in items:
            units_by_vm[key].append([fa, en])

    # Ecu_Menu: عمق مسیر پویاست — بعضی واحدها مستقیم آیتم دارن (۳ سطح مسیر:
    # برند/مدل/واحد) و بعضی‌ها قبل از آیتم‌ها یک سطح «زیرواحد» هم دارن
    # (۴ سطح: برند/مدل/واحد/زیرواحد) — مثل دلفی که یورو۲/یورو۴/یورو۵ و... داره.
    combo = {}
    units_meta = {}
    subunits_by_vmu = {}
    ws = wb["Ecu_Menu"]
    for r, c in _xl_find_anchors(ws):
        path_levels, items = _xl_read_block(ws, r, c)
        if len(path_levels) < 3:
            continue
        vfa, mfa = path_levels[0][0], path_levels[1][0]
        ufa, uen, uoid = path_levels[2]
        ukey = f"{vfa}||{mfa}||{ufa}"
        units_meta.setdefault(ukey, {"en": uen, "oldid": uoid})

        if len(path_levels) >= 4:
            sfa, sen, soid = path_levels[3]
        else:
            sfa, sen, soid = "", "", ""

        sub_list = subunits_by_vmu.setdefault(ukey, [])
        if [sfa, sen] not in sub_list:
            sub_list.append([sfa, sen])

        combo_key = f"{vfa}|{mfa}|{ufa}|{sfa}"
        combo[combo_key] = {
            "unit_en": uen,
            "unit_oldid": uoid,
            "sub_fa": sfa,
            "sub_en": sen,
            "sub_oldid": soid,
            "items": [[fa, en, oid] for fa, en, oid in items],
        }

    if not vehicles:
        raise ValueError("هیچ برندی توی شیت Vehicles پیدا نشد؛ ساختار فایل اکسل رو بررسی کنید.")

    return {
        "vehicles": vehicles,
        "models_by_veh": models_by_veh,
        "units_by_vm": units_by_vm,
        "units_meta": units_meta,
        "subunits_by_vmu": subunits_by_vmu,
        "combo": combo,
    }


class TreeData:
    def __init__(self, raw):
        self.vehicles = raw["vehicles"]
        self.models_by_veh = raw["models_by_veh"]
        self.units_by_vm = raw["units_by_vm"]
        self.units_meta = dict(raw.get("units_meta", {}))
        self.subunits_by_vmu = {k: [list(s) for s in v] for k, v in raw.get("subunits_by_vmu", {}).items()}
        self.combo = {}

        for key, rec in raw["combo"].items():
            parts = key.split("|")
            if len(parts) == 3:
                # سازگاری با فرمت قدیمی tree_data.json (بدون سطح زیرواحد):
                # با یک زیرواحدِ خالی (یعنی «نیازی به زیرواحد نیست») به فرمت
                # جدید ۴بخشی نگاشت می‌شه.
                vfa, mfa, ufa = parts
                sfa = ""
                new_key = f"{vfa}|{mfa}|{ufa}|{sfa}"
                new_rec = dict(rec)
                new_rec.setdefault("sub_fa", "")
                new_rec.setdefault("sub_en", "")
                new_rec.setdefault("sub_oldid", "")
                self.combo[new_key] = new_rec
                ukey = f"{vfa}||{mfa}||{ufa}"
                self.units_meta.setdefault(ukey, {"en": rec.get("unit_en", ""), "oldid": rec.get("unit_oldid", "")})
                self.subunits_by_vmu.setdefault(ukey, [["", ""]])
            else:
                self.combo[key] = rec

    @classmethod
    def from_json_file(cls, path):
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
        return cls(raw)

    @classmethod
    def from_excel_file(cls, path):
        return cls(build_tree_from_excel(path))

    def vehicle_fa_list(self):
        return [v[0] for v in self.vehicles]

    def vehicle_en(self, vfa):
        for fa, en in self.vehicles:
            if fa == vfa:
                return en or ""
        return ""

    def model_fa_list(self, vfa):
        return [m[0] for m in self.models_by_veh.get(vfa, [])]

    def model_en(self, vfa, mfa):
        for fa, en in self.models_by_veh.get(vfa, []):
            if fa == mfa:
                return en or ""
        return ""

    def unit_fa_list(self, vfa, mfa):
        key = f"{vfa}||{mfa}"
        return [u[0] for u in self.units_by_vm.get(key, [])]

    def unit_meta(self, vfa, mfa, ufa):
        return self.units_meta.get(f"{vfa}||{mfa}||{ufa}", {"en": "", "oldid": ""})

    def subunit_fa_list(self, vfa, mfa, ufa):
        """اگر این واحد چند زیرمجموعه/نسخه‌ی واقعی داشته باشه (مثل یورو۲/یورو۵
        برای دلفی)، اسامی‌شون رو برمی‌گردونه؛ در غیر این صورت لیست خالی
        (یعنی نیازی به انتخاب زیرواحد نیست)."""
        subs = self.subunits_by_vmu.get(f"{vfa}||{mfa}||{ufa}", [])
        return [s[0] for s in subs if s[0]]

    def combo_record(self, vfa, mfa, ufa, sfa=""):
        key = f"{vfa}|{mfa}|{ufa}|{sfa}"
        return self.combo.get(key)


class App(tk.Tk):
    CATEGORY_MOTORCYCLE = "موتورسیکلت"
    CATEGORY_CAR = "سواری"
    CATEGORY_SETTINGS = "تنظیمات"
    # ترتیب صفحه اول: ماشین → موتور → تنظیمات
    CATEGORY_OPTIONS = [CATEGORY_CAR, CATEGORY_MOTORCYCLE, CATEGORY_SETTINGS]
    # آیکونِ هر کارت: عکس آماده توی ui_icons
    CATEGORY_ICONS = {
        CATEGORY_MOTORCYCLE: ("image", "cat_motorcycle.png"),
        CATEGORY_CAR: ("image", "cat_car.png"),
        CATEGORY_SETTINGS: ("image", "settings_worker.png"),
    }

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.configure(bg=COLOR_BG)
        self.geometry("760x520")
        self.minsize(680, 460)
        self._set_app_icon()
        self._build_icon_index()

        # نشانگر آچار فقط اگر فایل .cur معتبر باشد؛ وگرنه hand2
        # (اعمال ناامنِ cursor باعث بالا نیامدن برنامه روی بعضی سیستم‌ها می‌شد)
        self.click_cursor = "hand2"
        resolved = self._resolve_click_cursor()
        if resolved != "hand2":
            try:
                self.configure(cursor=resolved)
                self.click_cursor = resolved
            except Exception:
                self.click_cursor = "hand2"
                try:
                    self.configure(cursor="hand2")
                except Exception:
                    pass
        else:
            try:
                self.configure(cursor="hand2")
            except Exception:
                pass

        init_fonts()  # اکنون که پنجره‌ی Tk ساخته شده، لیست فونت‌های سیستم قابل خواندن است

        self.config = load_config()
        # فقط از کش محلی که از گیت‌هاب همگام می‌شود خوانده می‌شود (نه پوشهٔ کاربر)
        self.diag_folder = diag_cache_folder()
        self.database_folder = database_cache_folder()
        self.motorcycle_data = None
        self._motorcycle_source_desc = "—"
        # حافظه اسکرول هر سطح منو؛ فقط موقع بازگشت بازیابی می‌شود
        self._scroll_memory = {}
        self._nav_direction = "forward"  # "forward" | "back"
        self.car_maker_data = {}    # کش: اسم شرکت -> TreeData (فقط وقتی انتخاب بشه خونده می‌شه)
        self.car_maker_source = {}  # اسم شرکت -> اسم فایل (برای نمایش کوچیک)
        self.data = None
        self._data_source_desc = "—"

        self._build_style()
        self._build_ui()

        # ابتدا از گیت‌هاب همگام‌سازی، بعد بارگذاری داده از کش
        self.after(400, self._startup_sync_and_load)

    def _startup_sync_and_load(self):
        """با باز شدن برنامه: بررسی سرور، در صورت نیاز پیام آپدیت، بعد بارگذاری داده."""
        def worker():
            try:
                self._sync_excels_from_github(parent=self, from_startup=True)
            except Exception:
                pass
            self.after(0, self._after_sync_load_data)
        threading.Thread(target=worker, daemon=True).start()

    def _after_sync_load_data(self):
        self.car_maker_data = {}
        self.car_maker_source = {}
        self._load_motorcycle_data()
        self._screen_refresh()

    def _sync_excels_from_github(self, parent=None, from_startup=False, silent_if_none=False, force_all=False):
        """
        بدون مانیفست: فایل‌های شناخته‌شده را با سرور مقایسه می‌کند.
        force_all=True → دانلود اجباری همهٔ فایل‌های موجود روی سرور.
        """
        parent = parent or self
        base = GITHUB_EXCEL_RAW_BASE
        token = None
        self.diag_folder = diag_cache_folder()
        self.database_folder = database_cache_folder()
        local_meta = self.config.get(CFG_LOCAL_EXCEL_META) or {}

        try:
            updates = find_updates_without_manifest(
                base, local_meta, self.diag_folder, token=token, force_all=force_all)
        except Exception as e:
            if not silent_if_none and not from_startup:
                self.after(0, lambda: messagebox.showwarning(
                    "آفلاین",
                    "اتصال به سرور برقرار نشد.\n"
                    "از آخرین نسخهٔ ذخیره‌شده استفاده می‌شود.\n\n"
                    f"{e}",
                    parent=parent))
            return False

        if not updates:
            if not silent_if_none and not from_startup:
                self.after(0, lambda: messagebox.showinfo(
                    "به‌روزرسانی",
                    "همهٔ فایل‌های اکسل با سرور یکسان هستند.\n"
                    "اگر مطمئنید فایل را عوض کرده‌اید، دکمهٔ «دانلود اجباری» را بزنید.",
                    parent=parent))
            return True

        names = "\n".join(f"  • {u['download_name']}" for u in updates)

        def ask_and_download():
            title = "دانلود اجباری از سرور" if force_all else "نیازمند به‌روزرسانی"
            msg = (
                f"{len(updates)} فایل از سرور دریافت می‌شود:\n\n{names}\n\n"
                "ادامه داده شود؟"
            )
            ok = messagebox.askyesno(title, msg, parent=parent)
            if not ok:
                return
            self._run_excel_download(updates, base, token, parent)

        self.after(0, ask_and_download)
        return True

    def _check_excel_updates(self, silent_if_none=False, parent=None, force_all=False):
        """دکمهٔ تنظیمات: بررسی / دانلود اجباری از سرور."""
        def worker():
            try:
                self._sync_excels_from_github(
                    parent=parent or self,
                    from_startup=False,
                    silent_if_none=silent_if_none,
                    force_all=force_all)
            except Exception:
                pass
        threading.Thread(target=worker, daemon=True).start()

    def _run_excel_download(self, updates, base, token, parent, on_done=None):
        """دانلود فایل‌ها با نوار پیشرفت ساده (پنجرهٔ modal)."""
        progress = tk.Toplevel(parent)
        progress.title("دانلود از سرور")
        progress.transient(parent)
        progress.grab_set()
        progress.geometry("420x140")
        progress.resizable(False, False)
        lbl = tk.Label(progress, text="در حال دانلود از گیت‌هاب…", font=(FONT_FA, 10), justify="right")
        lbl.pack(pady=(18, 6), padx=16)
        bar = ttk.Progressbar(progress, mode="determinate", maximum=max(len(updates), 1))
        bar.pack(fill="x", padx=20, pady=8)
        status = tk.Label(progress, text="", font=(FONT_FA, 9), fg="#555")
        status.pack(pady=4)

        def do_download():
            local_meta = dict(self.config.get(CFG_LOCAL_EXCEL_META) or {})
            errors = []
            for i, u in enumerate(updates):
                self.after(0, lambda i=i, n=u["download_name"]: (
                    bar.configure(value=i),
                    status.configure(text=f"دانلود: {n}"),
                    lbl.configure(text=f"فایل {i + 1} از {len(updates)}")))
                try:
                    sha = download_excel_file(
                        base, u["rel_path"], u["local_path"], token=token)
                    try:
                        size = os.path.getsize(u["local_path"])
                    except OSError:
                        size = u.get("remote_size")
                    local_meta[u["rel_path"]] = {
                        "etag": u.get("remote_etag") or "",
                        "size": size,
                        "sha256": sha or "",
                    }
                except Exception as e:
                    errors.append(f"{u['download_name']}: {e}")
            self.config[CFG_LOCAL_EXCEL_META] = local_meta
            save_config(self.config)
            success = not errors

            def finish():
                try:
                    progress.grab_release()
                    progress.destroy()
                except Exception:
                    pass
                if errors:
                    messagebox.showwarning(
                        "دانلود ناقص",
                        "برخی فایل‌ها از سرور دانلود نشدند:\n\n" + "\n".join(errors),
                        parent=parent)
                else:
                    messagebox.showinfo(
                        "موفق",
                        f"{len(updates)} فایل از سرور دریافت شد.\n"
                        "دادهٔ برنامه بارگذاری می‌شود.",
                        parent=parent)
                self.car_maker_data = {}
                self.car_maker_source = {}
                self._load_motorcycle_data()
                self.clear_all()
                if on_done:
                    try:
                        on_done(success)
                    except Exception:
                        pass

            self.after(0, finish)

        threading.Thread(target=do_download, daemon=True).start()

    def _resolve_click_cursor(self):
        """نشانگر کلیک. فعلاً همیشه hand2 تا از کرش احتمالی فایل .cur
        جلوگیری شود (قبلاً باعث بالا نیامدن برنامه می‌شد)."""
        return "hand2"

    def _set_app_icon(self):
        """آیکون برنامه (روی پنجره و نوار وظیفه‌ی ویندوز) رو تنظیم می‌کنه،
        اگه فایل icon.ico کنار برنامه/exe موجود باشه."""
        ico_path = resource_path("icon.ico")
        if os.path.exists(ico_path):
            try:
                self.iconbitmap(ico_path)
            except Exception:
                pass

    @staticmethod
    def _norm_name(name):
        return re.sub(r"[^a-z0-9]", "", (name or "").lower())

    def _build_icon_index(self):
        """پوشه‌ی brand_icons رو یک‌بار می‌خونه و بر اساس نام انگلیسیِ
        نرمال‌شده (بدون فاصله/حساسیت به بزرگی حروف) ایندکس می‌کنه، تا
        بشه لوگوی هر برند/مدل رو با نام انگلیسی‌اش توی اکسل پیدا کرد."""
        self._icon_files = {}
        self._icon_cache = {}
        folder = resource_path(ICONS_DIR)
        if os.path.isdir(folder):
            for fn in os.listdir(folder):
                stem, ext = os.path.splitext(fn)
                if ext.lower() not in (".png", ".jpg", ".jpeg"):
                    continue
                self._icon_files[self._norm_name(stem)] = os.path.join(folder, fn)

    def _load_icon(self, name_en, size=ICON_SIZE):
        """آیکون برند/مدل رو بر اساس نام انگلیسی‌اش و سایز دلخواه
        برمی‌گردونه؛ اگه عکسی براش پیدا نشه، None برمی‌گردونه (برای
        لوگوی بزرگِ کنار لیست) — استفاده‌کننده باید این حالت رو مدیریت
        کنه."""
        key = self._norm_name(name_en)
        cache_key = (key, size)
        if cache_key in self._icon_cache:
            return self._icon_cache[cache_key]
        path = self._icon_files.get(key)
        img = None
        if path:
            try:
                from PIL import Image, ImageTk
                im = Image.open(path).convert("RGBA")
                im.thumbnail((size, size), Image.LANCZOS)
                canvas = Image.new("RGBA", (size, size), (0, 0, 0, 0))
                canvas.paste(im, ((size - im.width) // 2, (size - im.height) // 2), im)
                img = ImageTk.PhotoImage(canvas)
            except Exception:
                img = None
        self._icon_cache[cache_key] = img
        return img

    def _load_motorcycle_data(self):
        """دیتای موتورسیکلت را فقط از کش دانلود‌شده از گیت‌هاب می‌خواند."""
        self.diag_folder = diag_cache_folder()
        path = os.path.join(self.diag_folder, MOTORCYCLE_EXCEL_NAME)
        if os.path.exists(path):
            try:
                self.motorcycle_data = TreeData.from_excel_file(path)
                self._motorcycle_source_desc = f"سرور / {MOTORCYCLE_EXCEL_NAME}"
                return True
            except Exception as e:
                messagebox.showwarning(
                    "خطا در خواندن اکسل موتورسیکلت",
                    f"{path}\n\n{e}"
                )
        self.motorcycle_data = None
        self._motorcycle_source_desc = "— (هنوز از سرور دانلود نشده)"
        return False

    def _load_car_maker(self, label):
        """دیتای یک شرکت خودروساز را فقط از کش گیت‌هاب می‌خواند."""
        if label in self.car_maker_data:
            return self.car_maker_data[label]
        filename = _CAR_LABEL_TO_FILE.get(label)
        if not filename:
            return None
        self.diag_folder = diag_cache_folder()
        path = os.path.join(self.diag_folder, filename)
        if not os.path.exists(path):
            messagebox.showwarning(
                "فایل پیدا نشد",
                f"فایل اکسل «{label}» روی این سیستم نیست.\n"
                "از تنظیمات «بررسی به‌روزرسانی» را بزنید تا از سرور دانلود شود.")
            return None
        try:
            data = TreeData.from_excel_file(path)
        except Exception as e:
            messagebox.showerror("خطا در خواندن اکسل", f"فایل «{label}» قابل خواندن نبود:\n{e}")
            return None
        self.car_maker_data[label] = data
        self.car_maker_source[label] = filename
        return data

    def _switch_category(self, value):
        """بر اساس دسته‌ی انتخاب‌شده، دیتای مربوطه رو فعال می‌کنه. اگه اون
        دیتا هنوز آماده نباشه، پیغام مناسب نشون می‌ده و False برمی‌گردونه
        (تا کاربر توی همون صفحه‌ی انتخاب بمونه)."""
        if value == self.CATEGORY_MOTORCYCLE:
            if self.motorcycle_data is None:
                messagebox.showwarning(
                    "داده‌ای نیست",
                    "دیتای موتورسیکلت از سرور دریافت نشده است.\n"
                    "اتصال اینترنت را بررسی کنید یا از تنظیمات «بررسی به‌روزرسانی» را بزنید.")
                return False
            self.data = self.motorcycle_data
            self._data_source_desc = self._motorcycle_source_desc
            if hasattr(self, "source_value_label"):
                self.source_value_label.config(text=self._data_source_desc)
            return True
        if value == self.CATEGORY_CAR:
            # فقط می‌ریم صفحه‌ی انتخاب شرکت خودروساز؛ دیتای واقعی وقتی ست
            # می‌شه که کاربر یکی از شرکت‌ها رو انتخاب کنه (چون هرکدوم فایل
            # اکسل جدا دارن).
            self.data = None
            return True
        # هیچ‌کدوم؛ تنظیمات خودش قبل از رسیدن به اینجا توی _screen_pick مدیریت می‌شه.
        return False

    def open_settings(self):
        """پاپ‌آپ تنظیمات مدرن (کارت مرکزی، سبک موبایل/iOS)."""
        if getattr(self, "_settings_overlay", None) is not None:
            try:
                if self._settings_overlay.winfo_exists():
                    return
            except Exception:
                pass

        self.update_idletasks()
        W = max(self.winfo_width(), 420)
        H = max(self.winfo_height(), 320)

        # لایه تیره نرم تمام‌صفحه
        overlay = tk.Frame(self, bg="#0f172a", cursor=self.click_cursor)
        overlay.place(x=0, y=0, relwidth=1, relheight=1)
        self._settings_overlay = overlay

        card_w = min(400, max(300, int(W * 0.72)))
        card_h = min(280, max(220, int(H * 0.5)))
        card_x = (W - card_w) // 2
        card_y = max(8, (H - card_h) // 2)

        # سایه ملایم پشت کارت
        shadow = tk.Frame(overlay, bg="#020617")
        shadow.place(x=card_x + 4, y=card_y + 6, width=card_w, height=card_h)

        # کارت اصلی
        card = tk.Frame(overlay, bg="#ffffff",
                         highlightbackground="#e2e8f0", highlightthickness=1)
        card.place(x=card_x, y=card_y, width=card_w, height=card_h)

        # هدر آبی با عنوان
        header = tk.Frame(card, bg=DEV_HEADER_BG, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)

        # آیکون کارگر/تنظیمات در هدر
        gear_img = self._load_ui_icon_sized("settings_worker.png", 32)
        if gear_img is not None:
            g_lbl = tk.Label(header, image=gear_img, bg=DEV_HEADER_BG)
            g_lbl.image = gear_img
            g_lbl.pack(side="right", padx=(0, 14), pady=12)
        tk.Label(header, text="تنظیمات", font=(FONT_FA, 15, "bold"),
                 bg=DEV_HEADER_BG, fg="#ffffff").pack(side="right", padx=(0, 6), pady=16)

        # دکمه بستن (X) سمت چپ هدر
        close_x = tk.Label(header, text="✕", font=(FONT_EN_NAME, 13, "bold"),
                            bg=DEV_HEADER_BG, fg="#c5d8f5", cursor=self.click_cursor)
        close_x.pack(side="left", padx=14, pady=14)

        body = tk.Frame(card, bg="#ffffff")
        body.pack(fill="both", expand=True, padx=20, pady=(18, 8))

        tk.Label(
            body,
            text="منبع داده فقط سرور گیت‌هاب است.\n"
                 "فایل‌های اکسل محلی دیگر استفاده نمی‌شوند.",
            font=(FONT_FA, 9), bg="#ffffff", fg="#64748b",
            justify="right", anchor="e").pack(fill="x", pady=(0, 14))

        def close_popup():
            try:
                overlay.destroy()
            except Exception:
                pass
            self._settings_overlay = None

        def on_check_updates():
            self._check_excel_updates(silent_if_none=False, parent=self, force_all=False)

        def on_force_download():
            self._check_excel_updates(silent_if_none=False, parent=self, force_all=True)

        check_btn = tk.Button(
            body, text="بررسی و دانلود از سرور",
            font=(FONT_FA, 11, "bold"),
            bg=DEV_HEADER_BG, fg="#ffffff",
            activebackground=DEV_HEADER_HOVER, activeforeground="#ffffff",
            bd=0, relief="flat", cursor=self.click_cursor,
            command=on_check_updates, pady=12)
        check_btn.pack(fill="x", pady=(4, 6))
        check_btn.bind("<Enter>", lambda e: check_btn.configure(bg=DEV_HEADER_HOVER))
        check_btn.bind("<Leave>", lambda e: check_btn.configure(bg=DEV_HEADER_BG))

        force_btn = tk.Button(
            body, text="دانلود اجباری همهٔ اکسل‌ها از سرور",
            font=(FONT_FA, 10, "bold"),
            bg="#64748b", fg="#ffffff",
            activebackground="#475569", activeforeground="#ffffff",
            bd=0, relief="flat", cursor=self.click_cursor,
            command=on_force_download, pady=10)
        force_btn.pack(fill="x", pady=(0, 8))

        tk.Label(
            body,
            text="اول اکسل را روی گیت‌هاب آپلود کنید،\nبعد این دکمه‌ها را بزنید.",
            font=(FONT_FA, 8), bg="#ffffff", fg="#94a3b8",
            justify="right", anchor="e").pack(fill="x")

        # دکمه پایین تمام‌عرض
        footer = tk.Frame(card, bg="#ffffff")
        footer.pack(fill="x", padx=20, pady=(4, 18))
        done_btn = tk.Button(
            footer, text="تأیید", font=(FONT_FA, 11, "bold"),
            bg=DEV_HEADER_BG, fg="#ffffff",
            activebackground=DEV_HEADER_HOVER, activeforeground="#ffffff",
            bd=0, relief="flat", cursor=self.click_cursor,
            command=close_popup, pady=10)
        done_btn.pack(fill="x")
        done_btn.bind("<Enter>", lambda e: done_btn.configure(bg=DEV_HEADER_HOVER))
        done_btn.bind("<Leave>", lambda e: done_btn.configure(bg=DEV_HEADER_BG))

        close_x.bind("<Button-1>", lambda e: close_popup())
        # کلیک روی پس‌زمینه تیره = بستن؛ کلیک روی کارت نباید حباب شود
        overlay.bind("<Button-1>", lambda e: close_popup())

        def _stop(e):
            return "break"

        card.bind("<Button-1>", _stop)
        header.bind("<Button-1>", _stop)
        body.bind("<Button-1>", _stop)
        footer.bind("<Button-1>", _stop)

        # ورود نرم: کمی از پایین + fade شبیه‌سازی با جابه‌جایی
        start_y = card_y + 28
        card.place(x=card_x, y=start_y, width=card_w, height=card_h)
        shadow.place(x=card_x + 4, y=start_y + 6, width=card_w, height=card_h)

        def _slide_in(step=0, steps=9):
            if step > steps:
                return
            t = 1.0 - (1.0 - step / steps) ** 3
            y = int(start_y + (card_y - start_y) * t)
            try:
                card.place(x=card_x, y=y, width=card_w, height=card_h)
                shadow.place(x=card_x + 4, y=y + 6, width=card_w, height=card_h)
            except Exception:
                return
            self.after(14, lambda: _slide_in(step + 1, steps))

        self.after(8, _slide_in)
        overlay.lift()

    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TCombobox", font=FONT_COMBO, padding=3)
        style.configure("Header.TLabel", font=FONT_HEADER,
                         background=COLOR_HEADER, anchor="center", padding=4)
        style.configure("TButton", font=FONT_BUTTON, padding=4)
        style.configure("Treeview", font=(FONT_FA, 9), rowheight=22)
        style.configure("Treeview.Heading", font=(FONT_FA, 9, "bold"))
        # فونت خود کرکره‌ی باز شونده (لیست گزینه‌ها) که به‌صورت جدا تنظیم می‌شود
        self.option_add("*TCombobox*Listbox.font", FONT_COMBO)
        self.option_add("*TCombobox*Listbox.selectBackground", "#CFE2F3")
        self.option_add("*TCombobox*Listbox.selectForeground", "#000000")

    def _build_ui(self):
        # متغیرهای وضعیتِ مسیر انتخاب (برند/مدل/واحد/زیرواحد) — دیگه به هیچ
        # کرکره‌ای وصل نیستن، فقط توسط «صفحه‌ی دستگاه» زیر مقداردهی می‌شن.
        self.veh_var = tk.StringVar()
        self.model_var = tk.StringVar()
        self.unit_var = tk.StringVar()
        self.sub_var = tk.StringVar()
        self.nav_path = []
        self.screen_level = "vehicle"
        self._current_content_widget = None
        self._database_excel_path = None

        # این نوار قبل از بدنه‌ی اصلی (بزل دستگاه) pack می‌شه تا فضای
        # ثابتِ خودش رو از پایین نگه داره؛ بزل با expand=True هرچی از
        # فضا باقی موند رو پر می‌کنه.
        self._build_bottom_shortcut_bar()
        self._build_device_screen()
        self._update_database_shortcut()

    def _build_bottom_shortcut_bar(self):
        """میان‌برهای همیشگی پایینِ برنامه: دو آیکون اکسل کنار هم —
        Database (دیتابیس) و Diag Menu (منوی دیاگ فعال). انتخاب مسیر
        فقط از تنظیمات انجام می‌شه؛ اینجا فقط باز کردن فایل."""
        bar = tk.Frame(self, bg=COLOR_BG)
        bar.pack(side="bottom", fill="x", padx=14, pady=(0, 8))

        excel_img = self._load_ui_icon_sized("excel_icon.png", 30)

        # --- Database ---
        db_holder = tk.Frame(bar, bg=COLOR_BG, cursor=self.click_cursor)
        db_holder.pack(side="left", padx=(0, 18))
        db_icon = tk.Label(db_holder, image=excel_img, bg=COLOR_BG, cursor=self.click_cursor)
        db_icon.image = excel_img
        db_icon.pack()
        db_caption = tk.Label(db_holder, text="Database", font=(FONT_EN_NAME, 7),
                               bg=COLOR_BG, fg="#555555", cursor=self.click_cursor)
        db_caption.pack()
        for w in (db_holder, db_icon, db_caption):
            w.bind("<Button-1>", lambda e: self._open_database_excel())

        # --- Diag Menu (اکسل منوی فعال) ---
        menu_holder = tk.Frame(bar, bg=COLOR_BG, cursor=self.click_cursor)
        menu_holder.pack(side="left")
        self._diag_menu_holder = menu_holder
        menu_icon = tk.Label(menu_holder, image=excel_img, bg=COLOR_BG, cursor=self.click_cursor)
        menu_icon.image = excel_img
        menu_icon.pack()
        self._diag_menu_icon = menu_icon
        menu_caption = tk.Label(menu_holder, text="Diag Menu", font=(FONT_EN_NAME, 7),
                                 bg=COLOR_BG, fg="#555555", cursor=self.click_cursor)
        menu_caption.pack()
        self._diag_menu_caption = menu_caption
        for w in (menu_holder, menu_icon, menu_caption):
            w.bind("<Button-1>", lambda e: self._open_current_excel())
        # ابتدا مخفی؛ با _update_excel_shortcut نشون داده می‌شه
        menu_holder.pack_forget()

    def _load_ui_icon(self, name):
        """آیکون‌های ثابت رابط کاربری (خانه/بازگشت/فلش‌های اسکرول) رو از
        پوشه‌ی ui_icons بارگذاری و کش می‌کنه."""
        cache_key = f"__ui__{name}"
        if cache_key in self._icon_cache:
            return self._icon_cache[cache_key]
        path = resource_path(os.path.join("ui_icons", name))
        img = None
        if os.path.exists(path):
            try:
                from PIL import Image, ImageTk
                img = ImageTk.PhotoImage(Image.open(path))
            except Exception:
                img = None
        self._icon_cache[cache_key] = img
        return img

    def _build_device_screen(self):
        """صفحه‌ی انتخاب شبیه صفحه‌نمایش یک دستگاه اسکنر واقعی: نوار بالای
        آبی (دکمه‌ی خانه‌ی گرد + عنوان/مسیر)، لیست سفیدِ گزینه‌های همون
        مرحله (با نقطه‌ی کوچیک کنار هرکدوم، دقیقاً شبیه لیست‌های دستگاه)،
        و نوار پایین آبی با دکمه‌ی بازگشتِ گرد وسط. وقتی وارد یک برند
        بشیم، لوگوی بزرگ همون برند سمت چپ لیست نشون داده می‌شه."""
        bezel = tk.Frame(self, bg=DEV_BEZEL)
        bezel.pack(fill="both", expand=True, padx=14, pady=(6, 12))

        screen = tk.Frame(bezel, bg=DEV_BODY_BG, highlightthickness=2,
                           highlightbackground=DEV_HEADER_BG, highlightcolor=DEV_HEADER_BG)
        screen.pack(fill="both", expand=True, padx=10, pady=10)

        # --- نوار بالا (آبی): دکمه‌ی خانه‌ی گرد + مسیر انتخاب ---
        top_bar = tk.Frame(screen, bg=DEV_HEADER_BG)
        top_bar.pack(side="top", fill="x")

        home_img = self._load_ui_icon("home_icon.png")
        home_btn = tk.Button(top_bar, image=home_img, command=self._screen_home,
                              bg=DEV_HEADER_BG, activebackground=DEV_HEADER_BG,
                              bd=0, highlightthickness=0, cursor=self.click_cursor)
        home_btn.image = home_img
        home_btn.pack(side="left", padx=10, pady=8)

        # نوار بالا فقط دکمه‌ی خانه + مسیر انتخاب داره (طبق خواسته‌ات، اسم
        # فایل اکسل از اینجا حذف شد؛ دکمه‌ی تنظیمات هم قبلاً به کارتِ
        # «تنظیمات» توی صفحه‌ی اصلی منتقل شده).
        self.breadcrumb_frame = tk.Frame(top_bar, bg=DEV_HEADER_BG)
        self.breadcrumb_frame.pack(side="right", fill="x", expand=True, padx=(10, 8), pady=6)

        # --- نوار پایین (آبی): فقط دکمه‌ی بازگشتِ گرد وسط ---
        # (میان‌برهای اکسل به نوار پایینِ کل برنامه منتقل شدن)
        bottom_bar = tk.Frame(screen, bg=DEV_HEADER_BG)
        bottom_bar.pack(side="bottom", fill="x")
        bottom_bar.grid_columnconfigure(0, weight=1)
        bottom_bar.grid_columnconfigure(1, weight=0)
        bottom_bar.grid_columnconfigure(2, weight=1)

        back_img = self._load_ui_icon("back_icon.png")
        self.screen_back_btn = tk.Button(bottom_bar, image=back_img, command=self._screen_back,
                                          bg=DEV_HEADER_BG, activebackground=DEV_HEADER_BG,
                                          bd=0, highlightthickness=0, cursor=self.click_cursor)
        self.screen_back_btn.image = back_img
        self.screen_back_btn.grid(row=0, column=1, pady=8)

        # --- بدنه: سمت چپ لوگوی بزرگ برند (وقتی داخل یک برندیم) + سمت راست لیست/جدول ---
        body = tk.Frame(screen, bg=DEV_BODY_BG)
        body.pack(side="top", fill="both", expand=True, padx=10, pady=8)

        self.brand_logo_panel = tk.Frame(body, bg=DEV_BODY_BG, width=300)
        self.brand_logo_label = tk.Label(self.brand_logo_panel, bg=DEV_BODY_BG)
        self.brand_logo_label.pack(expand=True)
        # brand_logo_panel فقط وقتی لازمه pack می‌شه (توی _screen_refresh)

        self.content_area = tk.Frame(body, bg=DEV_BODY_BG)
        self.content_area.pack(side="left", fill="both", expand=True)

        self._build_category_screen(self.content_area)
        self._build_screen_list(self.content_area)
        self._build_screen_table(self.content_area)

        self._screen_refresh()

    def _build_category_screen(self, parent):
        """صفحه‌ی اصلی/اولِ برنامه: چند کارتِ کوچیک و وسط‌چین برای انتخاب
        بین موتورسیکلت، سواری، و تنظیمات (بدون اینکه کل صفحه رو اشغال کنن)."""
        self.category_holder = tk.Frame(parent, bg=DEV_BODY_BG)
        self.category_grid = tk.Frame(self.category_holder, bg=DEV_BODY_BG)
        self.category_grid.pack(expand=True)

    def _load_ui_icon_sized(self, name, size):
        """مثل _load_ui_icon ولی با تغییر سایز به یه اندازه‌ی مشخص
        (برای آیکون‌های کارتِ صفحه‌ی اصلی که باید همه یه‌جور باشن)."""
        cache_key = f"__ui_sized__{name}__{size}"
        if cache_key in self._icon_cache:
            return self._icon_cache[cache_key]
        path = resource_path(os.path.join("ui_icons", name))
        img = None
        if os.path.exists(path):
            try:
                from PIL import Image, ImageTk
                im = Image.open(path).convert("RGBA")
                im.thumbnail((size, size), Image.LANCZOS)
                img = ImageTk.PhotoImage(im)
            except Exception:
                img = None
        self._icon_cache[cache_key] = img
        return img

    def _make_category_card(self, parent, label):
        """کارت گرد و مینیمال برای صفحه اول: پس‌زمینه سفید با گوشه‌های
        نرم، حاشیه خاکستری ظریف، و هایلایت آبی موقع هاور. کل کارت کلیک‌پذیره."""
        icon_kind, icon_val = self.CATEGORY_ICONS.get(label, ("emoji", "•"))
        W, H, R = 168, 168, 22  # عرض، ارتفاع، شعاع گوشه

        # قاب بیرونی هم‌رنگ پس‌زمینه تا گوشه‌های گرد دیده بشن
        outer = tk.Frame(parent, bg=DEV_BODY_BG, width=W, height=H, cursor=self.click_cursor)
        outer.pack_propagate(False)

        canvas = tk.Canvas(outer, width=W, height=H, bg=DEV_BODY_BG,
                            highlightthickness=0, bd=0, cursor=self.click_cursor)
        canvas.pack(fill="both", expand=True)

        def _rounded_rect(c, x1, y1, x2, y2, r, **kw):
            # نقاط تکراری برای smooth=True تا گوشه‌ها واقعاً گرد بشن
            pts = [
                x1 + r, y1, x1 + r, y1,
                x2 - r, y1, x2 - r, y1,
                x2, y1, x2, y1 + r,
                x2, y1 + r, x2, y2 - r,
                x2, y2 - r, x2, y2,
                x2 - r, y2, x2 - r, y2,
                x1 + r, y2, x1 + r, y2,
                x1, y2, x1, y2 - r,
                x1, y2 - r, x1, y1 + r,
                x1, y1 + r, x1, y1,
            ]
            return c.create_polygon(pts, smooth=True, **kw)

        # سایه خیلی ملایم (یک لایه پشت)
        _rounded_rect(canvas, 4, 5, W - 2, H - 1, R,
                       fill="#e8edf3", outline="", tags="card_shadow")

        # آیکون
        if icon_kind == "image":
            img = self._load_ui_icon_sized(icon_val, 56)
            icon_item = canvas.create_image(W // 2, 62, image=img, anchor="center")
            outer._card_img = img  # جلوگیری از GC
        else:
            icon_item = canvas.create_text(W // 2, 62, text=icon_val,
                                            font=(FONT_EN_NAME, 32), fill=DEV_TEXT)

        # متن
        text_item = canvas.create_text(
            W // 2, 128, text=bidi_fix(label),
            font=(FONT_FA, 13, "bold"), fill=DEV_TEXT, anchor="center")

        def _draw_border(color, width=1):
            canvas.delete("card_body")
            _rounded_rect(canvas, 2, 2, W - 4, H - 4, R,
                           fill="#ffffff", outline=color, width=width, tags="card_body")
            canvas.tag_raise(icon_item)
            canvas.tag_raise(text_item)

        _draw_border("#d5dbe3", 1)

        def on_enter(_e):
            _draw_border(DEV_HEADER_BG, 2)

        def on_leave(_e):
            _draw_border("#d5dbe3", 1)

        def on_click(_e):
            self._screen_pick(label)

        for w in (outer, canvas):
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)
            w.bind("<Button-1>", on_click)
        return outer

    def _populate_category_grid(self, options):
        for w in self.category_grid.winfo_children():
            w.destroy()
        # راست به چپ: اولین گزینه (سواری) سمت راست می‌نشیند
        for label in options:
            card = self._make_category_card(self.category_grid, label)
            card.pack(side="right", padx=14, pady=14)

    def _build_screen_list(self, parent):
        """لیست قابل‌اسکرولِ گزینه‌های هر مرحله؛ کاملاً دستی ساخته شده
        (نه ttk.Treeview) تا دقیقاً شبیه ردیف‌های لیست دستگاه واقعی باشه:
        پس‌زمینه‌ی سفید، خط جداکننده‌ی نازک، و هایلایت فیروزه‌ای موقع هاور/کلیک."""
        self.screen_list_holder = tk.Frame(parent, bg=DEV_BODY_BG)

        # ستون باریک سمت راست (فلش بالا، اسکرول‌بار، فلش پایین) رو قبل از
        # ناحیه‌ی لیست pack می‌کنیم تا همیشه عرض ثابتش رو حفظ کنه، حتی وقتی
        # عرض کل پنجره کم می‌شه (اگر لیست expand=True زودتر pack بشه،
        # موقع کوچیک‌شدن پنجره فضای این ستون رو می‌گیره و اسکرول ناپدید/کوچیک می‌شه).
        scroll_col = tk.Frame(self.screen_list_holder, bg=DEV_BODY_BG, width=42)
        scroll_col.pack(side="right", fill="y", padx=(6, 0))
        scroll_col.pack_propagate(False)

        list_area = tk.Frame(self.screen_list_holder, bg=DEV_BODY_BG)
        list_area.pack(side="left", fill="both", expand=True)

        self.screen_canvas = tk.Canvas(list_area, bg=DEV_BODY_BG, highlightthickness=0)
        self.screen_canvas.pack(side="left", fill="both", expand=True)

        self.screen_rows_frame = tk.Frame(self.screen_canvas, bg=DEV_BODY_BG)
        self._rows_window = self.screen_canvas.create_window((0, 0), window=self.screen_rows_frame, anchor="nw")

        def _on_rows_configure(event=None):
            bbox = self.screen_canvas.bbox("all")
            if bbox:
                self.screen_canvas.configure(scrollregion=bbox)

        def _on_canvas_configure(event):
            self.screen_canvas.itemconfig(self._rows_window, width=event.width)
            bbox = self.screen_canvas.bbox("all")
            if bbox:
                self.screen_canvas.configure(scrollregion=bbox)

        self.screen_rows_frame.bind("<Configure>", _on_rows_configure)
        self.screen_canvas.bind("<Configure>", _on_canvas_configure)

        def _mousewheel(event):
            # فقط وقتی لیست منو دیده می‌شود اسکرول کن
            if getattr(self, "screen_level", None) in (None, "category", "done"):
                return
            if event.num == 4 or getattr(event, "delta", 0) > 0:
                self.screen_canvas.yview_scroll(-3, "units")
            else:
                self.screen_canvas.yview_scroll(3, "units")
            return "break"

        self._list_mousewheel = _mousewheel
        for w in (self.screen_canvas, self.screen_rows_frame, list_area, scroll_col,
                  self.screen_list_holder):
            w.bind("<MouseWheel>", _mousewheel)
            w.bind("<Button-4>", _mousewheel)  # Linux up
            w.bind("<Button-5>", _mousewheel)  # Linux down

        # دکمه‌ی فلش بالا، اسکرول‌بار (وسط)، دکمه‌ی فلش پایین — دقیقاً
        # شبیه عکسی که فرستادی: اسکرول بین دو فلش.
        # اگر آیکون‌ها به هر دلیلی بارگذاری نشن، به‌جای خالی موندنِ دکمه از
        # متن فلش (▲/▼) استفاده می‌شه تا همیشه چیزی قابل‌دیدن و کلیک باشه.
        up_img = self._load_ui_icon("chevron_up.png")
        down_img = self._load_ui_icon("chevron_down.png")
        up_btn = tk.Button(scroll_col, image=up_img, text=("" if up_img else "▲"),
                            command=lambda: self.screen_canvas.yview_scroll(-3, "units"),
                            bg=DEV_BODY_BG, fg=DEV_HEADER_BG, font=(FONT_EN_NAME, 10, "bold"),
                            bd=0, highlightthickness=0, cursor=self.click_cursor)
        up_btn.image = up_img
        up_btn.pack(side="top", pady=(4, 6))
        down_btn = tk.Button(scroll_col, image=down_img, text=("" if down_img else "▼"),
                              command=lambda: self.screen_canvas.yview_scroll(3, "units"),
                              bg=DEV_BODY_BG, fg=DEV_HEADER_BG, font=(FONT_EN_NAME, 10, "bold"),
                              bd=0, highlightthickness=0, cursor=self.click_cursor)
        down_btn.image = down_img
        down_btn.pack(side="bottom", pady=(6, 4))

        # استایل مدرن و باریک برای اسکرول‌بار (بدون فلش‌های پیش‌فرض زشتِ
        # ویندوز، چون همین بالا/پایین رو خودمون با دکمه‌های آبی داریم)
        style = ttk.Style(self)
        style.element_create("Modern.Scrollbar.trough", "from", "clam")
        style.element_create("Modern.Scrollbar.thumb", "from", "clam")
        style.layout("Modern.Vertical.TScrollbar", [
            ("Modern.Scrollbar.trough", {"sticky": "ns", "children": [
                ("Modern.Scrollbar.thumb", {"expand": "1", "sticky": "nswe"})
            ]})
        ])
        style.configure("Modern.Vertical.TScrollbar", gripcount=0, width=10,
                         troughcolor="#eef2f7", background=DEV_HEADER_BG,
                         bordercolor="#eef2f7", relief="flat", borderwidth=0)
        style.map("Modern.Vertical.TScrollbar",
                   background=[("pressed", DEV_HEADER_HOVER), ("active", DEV_HEADER_HOVER)])

        self.screen_scrollbar = ttk.Scrollbar(scroll_col, orient="vertical",
                                               style="Modern.Vertical.TScrollbar",
                                               command=self.screen_canvas.yview)
        self.screen_scrollbar.pack(side="top", fill="y", expand=True, pady=2)
        self.screen_canvas.configure(yscrollcommand=self.screen_scrollbar.set)

    def _populate_screen_list(self, options):
        for w in self.screen_rows_frame.winfo_children():
            w.destroy()
        row_h = 46
        for opt in options:
            row = tk.Frame(self.screen_rows_frame, bg=DEV_BODY_BG, height=row_h)
            row.pack(fill="x")
            row.pack_propagate(False)

            # متن ردیف رو به تکه‌های فارسی/غیرفارسی می‌شکنیم و هر تکه رو با
            # فونت خودش (فارسی یا انگلیسی) جدا می‌سازیم؛ این‌طوری حتی وسط
            # یک عنوان فارسی هم اعداد/کلمات لاتین همیشه به شکل انگلیسیِ
            # عادی نمایش داده می‌شن، نه به سبک اعداد فارسی.
            text_holder = tk.Frame(row, bg=DEV_BODY_BG)
            text_holder.pack(side="right", fill="both", expand=True, padx=14)
            run_labels = []
            for run in split_text_runs(opt):
                run_font = FONT_SCREEN_ROW if is_persian_text(run) else FONT_SCREEN_ROW_EN
                rl = tk.Label(text_holder, text=run, font=run_font, bg=DEV_BODY_BG, fg=DEV_TEXT)
                rl.pack(side="right")
                run_labels.append(rl)

            sep = tk.Frame(self.screen_rows_frame, bg=DEV_ROW_BORDER, height=1)
            sep.pack(fill="x")

            def on_enter(e, r=row, th=text_holder, labels=run_labels):
                r.configure(bg=DEV_SELECT_BG)
                th.configure(bg=DEV_SELECT_BG)
                for l in labels:
                    l.configure(bg=DEV_SELECT_BG, fg=DEV_SELECT_FG)

            def on_leave(e, r=row, th=text_holder, labels=run_labels):
                r.configure(bg=DEV_BODY_BG)
                th.configure(bg=DEV_BODY_BG)
                for l in labels:
                    l.configure(bg=DEV_BODY_BG, fg=DEV_TEXT)

            def on_click(e, v=opt):
                self._screen_pick(v)

            mw = getattr(self, "_list_mousewheel", None)
            for widget in [row, text_holder] + run_labels:
                widget.bind("<Enter>", on_enter)
                widget.bind("<Leave>", on_leave)
                widget.bind("<Button-1>", on_click)
                if mw:
                    widget.bind("<MouseWheel>", mw)
                    widget.bind("<Button-4>", mw)
                    widget.bind("<Button-5>", mw)
            if mw:
                sep.bind("<MouseWheel>", mw)
                sep.bind("<Button-4>", mw)
                sep.bind("<Button-5>", mw)

        # به‌روزرسانی ناحیه اسکرول بعد از ساخت ردیف‌ها
        self.update_idletasks()
        bbox = self.screen_canvas.bbox("all")
        if bbox:
            self.screen_canvas.configure(scrollregion=bbox)

    def _build_screen_table(self, parent):
        self.screen_done_holder = tk.Frame(parent, bg=DEV_BODY_BG)

        style = ttk.Style()
        style.configure("ScreenHeader.TLabel", font=FONT_HEADER, background=DEV_HEADER_BG,
                         foreground=DEV_HEADER_TEXT, anchor="center", padding=4)
        style.configure("ScreenHeaderEN.TLabel", font=(FONT_EN_NAME, 10, "bold"),
                         background=DEV_HEADER_BG, foreground=DEV_HEADER_TEXT, anchor="center", padding=4)

        # اگه جدول توی صفحه جا نشه، قابل‌اسکرول باشه (با همون اسکرول‌بار
        # مدرن). اسکرول‌بار رو اول pack می‌کنیم تا عرضش همیشه ثابت بمونه.
        self.done_scrollbar = ttk.Scrollbar(self.screen_done_holder, orient="vertical",
                                             style="Modern.Vertical.TScrollbar")
        self.done_scrollbar.pack(side="right", fill="y", padx=(4, 0))

        self.done_canvas = tk.Canvas(self.screen_done_holder, bg=DEV_BODY_BG, highlightthickness=0)
        self.done_canvas.pack(side="left", fill="both", expand=True)
        self.done_scrollbar.config(command=self.done_canvas.yview)
        self.done_canvas.configure(yscrollcommand=self.done_scrollbar.set)

        table_holder = tk.Frame(self.done_canvas, bg=DEV_BODY_BG)
        self._done_window = self.done_canvas.create_window((0, 0), window=table_holder, anchor="nw")

        def _on_table_configure(event=None):
            self.done_canvas.configure(scrollregion=self.done_canvas.bbox("all"))

        def _on_done_canvas_configure(event):
            self.done_canvas.itemconfig(self._done_window, width=event.width)

        table_holder.bind("<Configure>", _on_table_configure)
        self.done_canvas.bind("<Configure>", _on_done_canvas_configure)

        def _done_mousewheel(event):
            self.done_canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

        self.done_canvas.bind("<MouseWheel>", _done_mousewheel)
        table_holder.bind("<MouseWheel>", _done_mousewheel)

        headers = ["ردیف", "متن فارسی", "English", "Old ID"]
        header_styles = ["ScreenHeader.TLabel", "ScreenHeader.TLabel",
                          "ScreenHeaderEN.TLabel", "ScreenHeaderEN.TLabel"]
        widths = [4, 16, 16, 30]
        for c, (h, sty, w) in enumerate(zip(headers, header_styles, widths)):
            ttk.Label(table_holder, text=h, style=sty, width=w).grid(
                row=0, column=c, sticky="nsew", padx=1, pady=1)

        self.row_widgets = []
        self._make_row(table_holder, 1, COLOR_VEHICLE, bold=True)
        self._make_row(table_holder, 2, COLOR_MODEL, bold=True)
        self._make_row(table_holder, 3, COLOR_UNIT, bold=True)
        self._make_row(table_holder, 4, COLOR_SUBUNIT, bold=True)
        for i in range(ITEM_ROWS):
            self._make_row(table_holder, 5 + i, COLOR_ITEM, bold=False)

        col_weights = [1, 4, 4, 7]
        for c, wgt in enumerate(col_weights):
            table_holder.columnconfigure(c, weight=wgt)
        # screen_done_holder فقط وقتی لازمه pack می‌شه (توی _screen_refresh)

    def _resolve_screen(self):
        """بر اساس مسیر انتخاب‌شده‌ی فعلی (self.nav_path)، مشخص می‌کنه که
        صفحه الان باید چه سطحی رو نشون بده و گزینه‌های اون سطح چی هستن.
        سطح اول همیشه صفحه‌ی اصلی (کارت‌های فلش‌ایسیو/EOBD/موتورسیکلت/
        سواری/کاربری‌خاص). مسیر «سواری» چون هر شرکت خودروساز فایل اکسل
        جدا داره، یه سطحِ اضافه («شرکت خودروساز») قبل از برند/مدل داره؛
        مسیر «موتورسیکلت» دقیقاً مثل قبل یک‌فایلیه."""
        n = len(self.nav_path)
        if n == 0:
            return "category", self.CATEGORY_OPTIONS

        if self.nav_path[0] == self.CATEGORY_CAR:
            if n == 1:
                return "maker", CAR_MAKER_LABELS
            if n == 2:
                return "vehicle", (self.data.vehicle_fa_list() if self.data else [])
            if n == 3:
                return "model", self.data.model_fa_list(self.nav_path[2])
            if n == 4:
                return "unit", self.data.unit_fa_list(self.nav_path[2], self.nav_path[3])
            if n == 5:
                subs = self.data.subunit_fa_list(self.nav_path[2], self.nav_path[3], self.nav_path[4])
                if len(subs) > 1:
                    return "subunit", subs
                return "done", []
            return "done", []

        # موتورسیکلت (تک‌فایلی، مثل قبل)
        if n == 1:
            return "vehicle", (self.data.vehicle_fa_list() if self.data else [])
        if n == 2:
            return "model", self.data.model_fa_list(self.nav_path[1])
        if n == 3:
            return "unit", self.data.unit_fa_list(self.nav_path[1], self.nav_path[2])
        if n == 4:
            subs = self.data.subunit_fa_list(self.nav_path[1], self.nav_path[2], self.nav_path[3])
            if len(subs) > 1:
                return "subunit", subs
            return "done", []
        return "done", []

    LEVEL_TITLES = {
        "category": "",
        "maker": "شرکت خودروساز را انتخاب کنید",
        "vehicle": "برند را انتخاب کنید",
        "model": "مدل را انتخاب کنید",
        "unit": "نام ایسیو را انتخاب کنید",
        "subunit": "زیر واحد را انتخاب کنید",
    }

    # ردیف‌های برند (index 0، رنگ قرمز) و مدل (index 1، رنگ آبی) توی جدول
    # نهایی نیازی به نمایش ندارن (توی مسیر breadcrumb بالای صفحه از قبل
    # نشون داده می‌شن)، پس همیشه مخفی می‌مونن.
    HIDDEN_ROW_INDEXES = {0, 1}

    def _current_excel_path(self):
        """مسیر کاملِ فایل اکسلی که الان داره داده‌ی صفحه‌ی فعلی رو تأمین
        می‌کنه برمی‌گردونه (یا None اگه هنوز مشخص نیست، مثلاً توی صفحه‌ی
        اصلی یا موقع انتخاب شرکتِ خودروساز)."""
        if not self.nav_path:
            return None
        if self.nav_path[0] == self.CATEGORY_MOTORCYCLE:
            path = os.path.join(self.diag_folder, MOTORCYCLE_EXCEL_NAME)
            return path if os.path.exists(path) else None
        if self.nav_path[0] == self.CATEGORY_CAR and len(self.nav_path) >= 2:
            label = self.nav_path[1]
            filename = self.car_maker_source.get(label) or _CAR_LABEL_TO_FILE.get(label)
            if filename:
                path = os.path.join(self.diag_folder, filename)
                return path if os.path.exists(path) else None
        return None

    def _update_excel_shortcut(self):
        """میان‌بر Diag Menu پایین برنامه رو بر اساس اینکه الان فایل
        اکسل منوی مشخصی فعاله یا نه، نشون/مخفی می‌کنه."""
        path = self._current_excel_path()
        holder = getattr(self, "_diag_menu_holder", None)
        if path:
            self._active_excel_path = path
            if holder is not None:
                holder.pack(side="left")
        else:
            self._active_excel_path = None
            if holder is not None:
                holder.pack_forget()

    def _open_file_external(self, path):
        """یک فایل رو با برنامه‌ی پیش‌فرضِ سیستم (مثلاً اکسل) باز می‌کنه."""
        if not path or not os.path.exists(path):
            messagebox.showwarning("پیدا نشد", f"فایل پیدا نشد:\n{path}")
            return
        try:
            if sys.platform == "win32":
                os.startfile(path)  # noqa
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", path])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("خطا در باز کردن فایل", f"{path}\n\n{e}")

    def _open_current_excel(self):
        path = getattr(self, "_active_excel_path", None) or self._current_excel_path()
        if not path:
            messagebox.showinfo(
                "Diag Menu",
                "هنوز منوی دیاگی انتخاب نشده.\n"
                "وارد موتورسیکلت یا یکی از شرکت‌های سواری بشید تا فایل اکسل مربوطه فعال بشه.")
            return
        self._open_file_external(path)

    def _find_database_excel(self):
        """توی پوشه‌ی دیتابیس (Diag_Database) دنبال تنها فایل اکسل داخلش
        می‌گرده (فایل‌های قفلِ موقتِ اکسل که با ~$ شروع می‌شن نادیده گرفته
        می‌شن). اگه پوشه یا فایلی پیدا نشه، None برمی‌گردونه."""
        folder = self.database_folder
        if not folder or not os.path.isdir(folder):
            return None
        for fn in sorted(os.listdir(folder)):
            if fn.startswith("~$"):
                continue
            if os.path.splitext(fn)[1].lower() in (".xlsx", ".xlsm", ".xls"):
                return os.path.join(folder, fn)
        return None

    def _update_database_shortcut(self):
        """میان‌بر اکسلِ دیتابیس (پایینِ برنامه) رو به‌روزرسانی می‌کنه."""
        self._database_excel_path = self._find_database_excel()

    def _open_database_excel(self):
        path = getattr(self, "_database_excel_path", None) or self._find_database_excel()
        if not path:
            messagebox.showwarning(
                "پیدا نشد",
                f"هیچ فایل اکسلی توی پوشه‌ی دیتابیس پیدا نشد:\n{self.database_folder}\n\n"
                "می‌تونید مسیرش رو از کارتِ «تنظیمات» تغییر بدید.")
            return
        self._open_file_external(path)

    def _screen_refresh(self):
        level, options = self._resolve_screen()
        self.screen_level = level

        self._update_breadcrumb(level)
        self.screen_back_btn.config(state=("normal" if self.nav_path else "disabled"))

        # لوگوی بزرگ برند فقط وقتی داخل یک برند/مدل هستیم (نه توی صفحه‌ی
        # اصلی، انتخاب شرکت خودروساز، یا جدول نتیجه) سمت چپ لیست نشون
        # داده می‌شه. توی مسیر «سواری» برند یک سطح عمیق‌تره (بعد از شرکت).
        big_icon = None
        if self.data is not None and level != "done":
            is_car = bool(self.nav_path) and self.nav_path[0] == self.CATEGORY_CAR
            veh_idx = 2 if is_car else 1
            if len(self.nav_path) > veh_idx:
                vfa = self.nav_path[veh_idx]
                ven = self.data.vehicle_en(vfa)
                big_icon = self._load_icon(ven, size=280)
        if big_icon is not None:
            self.brand_logo_label.configure(image=big_icon)
            self.brand_logo_label.image = big_icon
            # «before=» لازمه چون content_area قبلاً pack شده؛ بدون این،
            # پنل لوگو به‌جای سمت چپِ واقعی، بعد از content_area (یعنی
            # عملاً وسط/سمت راست) قرار می‌گرفت.
            self.brand_logo_panel.pack(side="left", fill="y", padx=(0, 10), before=self.content_area)
        else:
            self.brand_logo_panel.pack_forget()

        if level == "category":
            new_widget = self.category_holder
            self._populate_category_grid(options)
        elif level == "done":
            new_widget = self.screen_done_holder
        else:
            new_widget = self.screen_list_holder
            self._populate_screen_list(options)

        # اسکرول و قاب صفحه ثابت می‌مانند؛ فقط محتوا جابه‌جا می‌شود (بدون
        # جابه‌جایی کل ویجت که اسکرول‌بار را هم حرکت دهد).
        for w in (self.category_holder, self.screen_list_holder, self.screen_done_holder):
            if w is not new_widget:
                w.pack_forget()
                try:
                    w.place_forget()
                except Exception:
                    pass
        new_widget.pack(fill="both", expand=True)
        self._current_content_widget = new_widget

        # اسکرول: حرکت رو به جلو از اول؛ بازگشت به جای قبلی
        if level not in ("category", "done"):
            going_back = getattr(self, "_nav_direction", "forward") == "back"
            if going_back:
                saved = self._scroll_memory.get(tuple(self.nav_path))
                def _restore(pos=saved):
                    self.update_idletasks()
                    bbox = self.screen_canvas.bbox("all")
                    if bbox:
                        self.screen_canvas.configure(scrollregion=bbox)
                    if pos is not None:
                        self.screen_canvas.yview_moveto(pos)
                    else:
                        self.screen_canvas.yview_moveto(0)
                self.after_idle(_restore)
            else:
                def _to_top():
                    self.update_idletasks()
                    bbox = self.screen_canvas.bbox("all")
                    if bbox:
                        self.screen_canvas.configure(scrollregion=bbox)
                    self.screen_canvas.yview_moveto(0)
                self.after_idle(_to_top)

        # افکت فقط روی نوشته‌ها / ردیف‌ها (نه روی اسکرول) — فقط رو به جلو
        if level == "done":
            self._animate_table_shutter()
        elif level not in ("category",) and getattr(self, "_nav_direction", "forward") == "forward":
            self._animate_list_text_rows()

        self._nav_direction = "forward"  # پیش‌فرض برای بعدی
        self._update_excel_shortcut()

    def _animate_list_text_rows(self):
        """افکت ظاهر شدن نوشته‌های لیست از بالا به پایین (اسکرول ثابت می‌ماند)."""
        if getattr(self, "_text_anim_ids", None):
            for aid in self._text_anim_ids:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
        self._text_anim_ids = []

        rows = []
        try:
            children = list(self.screen_rows_frame.winfo_children())
        except Exception:
            return
        # هر ردیف + جداکننده؛ فقط فریم‌های ردیف (ارتفاع‌دار) را انیمیت می‌کنیم
        for w in children:
            try:
                if int(w.cget("height") or 0) >= 20:
                    rows.append(w)
            except Exception:
                continue

        # ابتدا متن‌ها را موقتاً مخفی کن (fg هم‌رنگ پس‌زمینه)
        for row in rows:
            for child in row.winfo_children():
                self._set_text_fg_recursive(child, DEV_BODY_BG)

        def reveal(i):
            if i >= len(rows):
                return
            row = rows[i]
            for child in row.winfo_children():
                self._set_text_fg_recursive(child, DEV_TEXT)
            aid = self.after(28, lambda: reveal(i + 1))
            self._text_anim_ids.append(aid)

        if rows:
            self.after(40, lambda: reveal(0))

    def _set_text_fg_recursive(self, widget, color):
        try:
            if isinstance(widget, tk.Label):
                # موقع هاور ممکن است رنگ عوض شود؛ فقط اگر انتخاب‌شده نیست
                bg = widget.cget("bg")
                if bg == DEV_SELECT_BG:
                    widget.configure(fg=DEV_SELECT_FG if color != DEV_BODY_BG else DEV_SELECT_BG)
                else:
                    widget.configure(fg=color)
            for ch in widget.winfo_children():
                self._set_text_fg_recursive(ch, color)
        except Exception:
            pass

    def _animate_table_shutter(self):
        """جدول نهایی با افکت کرکره‌ای از بالا به پایین: هر ردیف منطقی
        یکی‌یکی ظاهر می‌شود. اسکرول‌بار و قاب ثابت می‌مانند."""
        if getattr(self, "_table_anim_id", None):
            try:
                self.after_cancel(self._table_anim_id)
            except Exception:
                pass
            self._table_anim_id = None

        try:
            self.done_canvas.yview_moveto(0)
        except Exception:
            pass

        # فقط ردیف‌هایی که الان در جدول grid شده‌اند (نه برند/مدل مخفی/خالی)
        row_groups = []
        for w in getattr(self, "row_widgets", []):
            entries = w.get("entries", [])
            if not entries:
                continue
            try:
                if not entries[0].grid_info():
                    continue  # الان مخفی است
            except Exception:
                continue
            for e in entries:
                try:
                    e.grid_remove()
                except Exception:
                    pass
            row_groups.append(entries)

        delay = 45  # میلی‌ثانیه بین هر ردیف

        def show_next(i):
            if i >= len(row_groups):
                self._table_anim_id = None
                return
            for e in row_groups[i]:
                try:
                    e.grid()
                except Exception:
                    pass
            self._table_anim_id = self.after(delay, lambda: show_next(i + 1))

        if row_groups:
            self.after(40, lambda: show_next(0))

    def _save_scroll_position(self):
        """موقعیت اسکرول لیست فعلی را برای بازگشت بعدی ذخیره می‌کند."""
        if getattr(self, "screen_level", None) in (None, "category", "done"):
            return
        if not hasattr(self, "screen_canvas"):
            return
        try:
            self._scroll_memory[tuple(self.nav_path)] = self.screen_canvas.yview()[0]
        except Exception:
            pass

    def _update_breadcrumb(self, level):
        for w in self.breadcrumb_frame.winfo_children():
            w.destroy()
        if not self.nav_path:
            tk.Label(self.breadcrumb_frame, text=self.LEVEL_TITLES.get(level, ""),
                     font=FONT_SCREEN_ROW, bg=DEV_HEADER_BG, fg=DEV_HEADER_TEXT_DIM).pack(side="right")
            return
        last = len(self.nav_path) - 1
        for i, seg in enumerate(self.nav_path):
            seg_color = DEV_HEADER_TEXT if i == last else DEV_HEADER_TEXT_DIM
            for run in split_text_runs(seg):
                run_font = FONT_SCREEN_ROW if is_persian_text(run) else FONT_SCREEN_ROW_EN
                tk.Label(self.breadcrumb_frame, text=run, font=run_font, bg=DEV_HEADER_BG,
                         fg=seg_color).pack(side="right")
            if i != last:
                tk.Label(self.breadcrumb_frame, text=" < ", font=FONT_SCREEN_ROW,
                         bg=DEV_HEADER_BG, fg=DEV_HEADER_TEXT_DIM).pack(side="right")

    def _screen_pick(self, value):
        self._save_scroll_position()
        self._nav_direction = "forward"
        if not self.nav_path:
            # این انتخابِ سطح اول (صفحه‌ی اصلی) هست
            if value == self.CATEGORY_SETTINGS:
                self.open_settings()
                return
            if not self._switch_category(value):
                return  # دیتای این دسته آماده نیست؛ توی همین صفحه می‌مونیم
            self.nav_path.append(value)
            self._screen_refresh()
            return

        is_car = self.nav_path[0] == self.CATEGORY_CAR
        if is_car and len(self.nav_path) == 1:
            # این انتخابِ شرکت خودروساز هست (سطح مخصوص مسیر «سواری»)
            maker_data = self._load_car_maker(value)
            if maker_data is None:
                return
            self.data = maker_data
            self._data_source_desc = self.car_maker_source.get(value, value)
            if hasattr(self, "source_value_label"):
                self.source_value_label.config(text=self._data_source_desc)
            self.nav_path.append(value)
            self._screen_refresh()
            return

        self.nav_path.append(value)
        n = len(self.nav_path)
        base = 2 if is_car else 1  # ایندکسی که «برند/مدل سطح اول» (vehicle) توش قرار می‌گیره
        if n == base + 1:
            self.veh_var.set(value)
            self.on_vehicle_change()
        elif n == base + 2:
            self.model_var.set(value)
            self.on_model_change()
        elif n == base + 3:
            self.unit_var.set(value)
            self.on_unit_change()
        elif n == base + 4:
            self.sub_var.set(value)
            self.on_subunit_change()
        self._screen_refresh()

    def _screen_back(self):
        if not self.nav_path:
            return
        self._save_scroll_position()
        self._nav_direction = "back"
        is_car = self.nav_path[0] == self.CATEGORY_CAR
        self.nav_path.pop()
        n = len(self.nav_path)
        base = 2 if is_car else 1
        if n <= base:
            self.veh_var.set("")
            self.model_var.set("")
            self.unit_var.set("")
            self.sub_var.set("")
            self._clear_from(0)
            if is_car and n <= 1:
                self.data = None  # هنوز شرکتی انتخاب نشده
        elif n == base + 1:
            self.model_var.set("")
            self.unit_var.set("")
            self.sub_var.set("")
            self._clear_from(1)
        elif n == base + 2:
            self.unit_var.set("")
            self.sub_var.set("")
            self._clear_from(2)
        elif n == base + 3:
            self.sub_var.set("")
            self._clear_from(3)
        self._screen_refresh()

    def _screen_home(self):
        # دکمه‌ی خانه همیشه برمی‌گرده به همون صفحه‌ی اصلی (انتخاب
        # موتورسیکلت/سواری/تنظیمات) — نه فقط یک سطح بالاتر.
        self._save_scroll_position()
        self._nav_direction = "forward"
        self.nav_path = []
        self.veh_var.set("")
        self.model_var.set("")
        self.unit_var.set("")
        self.sub_var.set("")
        self._clear_from(0)
        self._screen_refresh()

    def _make_cell(self, parent, r, c, color, bold, width, english=False):
        var = tk.StringVar()
        if english:
            font = FONT_CELL_EN_BOLD if bold else FONT_CELL_EN
        else:
            font = FONT_CELL_BOLD if bold else FONT_CELL
        entry = tk.Entry(parent, textvariable=var, justify="center", font=font,
                          relief="flat", readonlybackground=color, disabledbackground=color,
                          state="readonly", width=width, bd=1)
        entry.grid(row=r, column=c, sticky="nsew", padx=1, pady=1, ipady=3)
        return var, entry

    def _make_row(self, parent, r, color, bold):
        redif_var, redif_e = self._make_cell(parent, r, 0, color, bold, 4)
        fa_var, fa_e = self._make_cell(parent, r, 1, color, bold, 16)
        en_var, en_e = self._make_cell(parent, r, 2, color, bold, 16, english=True)
        oid_var, oid_e = self._make_cell(parent, r, 3, color, bold, 30, english=True)
        self.row_widgets.append({"redif": redif_var, "fa": fa_var, "en": en_var, "oldid": oid_var,
                                  "entries": [redif_e, fa_e, en_e, oid_e],
                                  "fa_entry": fa_e, "fa_bold": bold})

    def on_vehicle_change(self, event=None):
        vfa = self.veh_var.get()
        self.model_var.set("")
        self.unit_var.set("")
        self.sub_var.set("")
        self._clear_from(1)
        self._set_row(0, "", vfa, self.data.vehicle_en(vfa), "")

    def on_model_change(self, event=None):
        vfa = self.veh_var.get()
        mfa = self.model_var.get()
        self.unit_var.set("")
        self.sub_var.set("")
        self._clear_from(2)
        self._set_row(1, "", mfa, self.data.model_en(vfa, mfa), "")

    def on_unit_change(self, event=None):
        vfa = self.veh_var.get()
        mfa = self.model_var.get()
        ufa = self.unit_var.get()
        self._clear_from(3)
        meta = self.data.unit_meta(vfa, mfa, ufa)
        self._set_row(2, "", ufa, meta.get("en", ""), meta.get("oldid", ""))

        subs = self.data.subunit_fa_list(vfa, mfa, ufa)
        if not subs:
            # این واحد زیرمجموعه‌ای نداره؛ مستقیم آیتم‌ها رو نشون می‌دیم.
            self.sub_var.set("")
            self._fill_items(vfa, mfa, ufa, "")
        elif len(subs) == 1:
            # فقط یک زیرمجموعه هست؛ نیازی به انتخاب دستی نیست، خودکار
            # انتخاب و نمایش داده می‌شه.
            self.sub_var.set(subs[0])
            self.on_subunit_change()
        else:
            # چند زیرمجموعه/نسخه‌ی واقعی هست (مثل یورو۲/یورو۵ برای دلفی)؛
            # صفحه‌ی دستگاه لیست‌شون رو نشون می‌ده تا کاربر انتخاب کنه.
            self.sub_var.set("")
            self._clear_from(4)

    def on_subunit_change(self, event=None):
        vfa = self.veh_var.get()
        mfa = self.model_var.get()
        ufa = self.unit_var.get()
        sfa = self.sub_var.get()
        rec = self.data.combo_record(vfa, mfa, ufa, sfa)
        if not rec:
            self._clear_from(3)
            self._warn_missing_combo(vfa, mfa, ufa, sfa)
            return
        self._fill_items(vfa, mfa, ufa, sfa, rec)

    def _fill_items(self, vfa, mfa, ufa, sfa, rec=None):
        if rec is None:
            rec = self.data.combo_record(vfa, mfa, ufa, sfa)
        if not rec:
            self._clear_from(3)
            self._warn_missing_combo(vfa, mfa, ufa, sfa)
            return
        self._set_row(3, "", rec.get("sub_fa", ""), rec.get("sub_en", ""), rec.get("sub_oldid", ""))
        items = rec.get("items", [])
        for i in range(ITEM_ROWS):
            row_idx = 4 + i
            if i < len(items):
                fa, en, oid = items[i]
                self._set_row(row_idx, str(i + 1), fa, en or "", oid or "")
            else:
                self._set_row(row_idx, "", "", "", "")

    def _warn_missing_combo(self, vfa, mfa, ufa, sfa):
        """وقتی برای برند/مدل/واحد انتخاب‌شده هیچ داده‌ای توی شیت Ecu_Menu
        پیدا نشه (معمولاً به‌خاطر ناهماهنگی/غلط املایی نام برند یا مدل
        بین شیت‌های مختلف اکسل)، به‌جای جدول خالیِ بی‌توضیح، یک پیغام
        روشن نشون می‌ده تا کاربر بتونه مشکل رو توی خود اکسل پیدا کنه."""
        path_txt = " > ".join([p for p in [vfa, mfa, ufa, sfa] if p])
        messagebox.showwarning(
            "داده‌ای پیدا نشد",
            f"برای «{path_txt}» هیچ ردیفی توی شیت Ecu_Menu فایل اکسل پیدا نشد.\n\n"
            "معمولاً دلیلش اینه که نام برند/مدل/واحد توی شیت Ecu_Menu با نام‌شون "
            "توی شیت‌های Vehicles/Option/Unit یکی نیست (مثلاً یک غلط املایی کوچیک).\n"
            "لطفاً توی اکسل، املای این مسیر رو توی همه‌ی شیت‌ها با هم مقایسه و یکسان کنید، "
            "بعد دکمه‌ی «🔄 بارگذاری مجدد داده‌ها» رو بزنید.")

    def _set_row(self, idx, redif, fa, en, oldid):
        w = self.row_widgets[idx]
        w["redif"].set(bidi_fix(redif))
        w["fa"].set(bidi_fix(fa))
        w["en"].set(bidi_fix(en))
        w["oldid"].set(bidi_fix(oldid))
        # اگه متنِ ستونِ «متن فارسی» این ردیف در واقع کاملاً انگلیسی باشه
        # (مثل «MSE60»)، فونتش رو موقتاً به فونت انگلیسی عوض می‌کنیم تا
        # عددهاش هم به شکل غربی (نه فارسی) نشون داده بشن.
        fa_bold = w.get("fa_bold", False)
        fa_font = (FONT_CELL_BOLD if fa_bold else FONT_CELL) if is_persian_text(fa) else \
                  (FONT_CELL_EN_BOLD if fa_bold else FONT_CELL_EN)
        w["fa_entry"].configure(font=fa_font)
        # ردیف‌های برند و مدل (index 0 و 1، همون قرمز/آبی) توی جدول نهایی
        # اصلاً لازم نیست نشون داده بشن؛ داده‌شون همچنان نگه داشته می‌شه
        # (چون جاهای دیگه‌ی کد بهش رجوع می‌کنه) ولی مخفی می‌مونه.
        # ردیف‌های کاملاً خالی (مثلاً وقتی آیتم‌های کمتری از ITEM_ROWS وجود
        # داره، یا زیرواحدی در کار نیست) هم اصلاً نشون داده نشن.
        is_empty = not (redif or fa or en or oldid)
        hide = is_empty or idx in self.HIDDEN_ROW_INDEXES
        for entry in w["entries"]:
            if hide:
                entry.grid_remove()
            else:
                entry.grid()

    def _clear_from(self, start_idx):
        for idx in range(start_idx, len(self.row_widgets)):
            self._set_row(idx, "", "", "", "")

    def clear_all(self):
        self.nav_path = []
        self.veh_var.set("")
        self.model_var.set("")
        self.unit_var.set("")
        self.sub_var.set("")
        self._clear_from(0)
        self._screen_refresh()


if __name__ == "__main__":
    app = App()
    app.mainloop()
